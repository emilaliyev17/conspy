from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ImproperlyConfigured
from django.utils.timezone import make_naive
from django.db.models import Q, Sum, Min
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils.text import slugify
from .models import (
    Company,
    FinancialData,
    ChartOfAccounts,
    DataBackup,
    CFDashboardMetric,
    CFDashboardData,
    ActiveState,
    SalaryData,
    PLComment,
    PLCommentFile,
    HubSpotData,
    HubSpotSyncLog,
)
from .forms import ActiveStateForm
from .feature_flags import is_enabled
from .services.hubspot_service import HubSpotService
import pandas as pd
import csv
from datetime import datetime, date
import json
import re
from decimal import Decimal, InvalidOperation
import logging
from dateutil.relativedelta import relativedelta
import calendar
import openpyxl.styles
import copy

logger = logging.getLogger(__name__)

def clean_number_value(value):
    """Clean and parse number values from Excel, handling various formats including QuickBooks."""
    if value is None or pd.isna(value):
        return None
    
    value_str = str(value).strip()
    if not value_str or value_str == '':
        return None
    
    # Remove quotes and apostrophes
    value_str = value_str.strip("'\"")
    
    # Remove currency symbols ($, €, £, etc.)
    value_str = value_str.replace('$', '').replace('€', '').replace('£', '').replace('¥', '')
    
    # Remove all spaces
    value_str = value_str.replace(' ', '')
    
    # Handle negative numbers in parentheses: (1,234.56) -> -1234.56
    if value_str.startswith('(') and value_str.endswith(')'):
        value_str = '-' + value_str[1:-1]
    
    # Handle QuickBooks formats with different separators
    # Remove thousand separators (commas)
    value_str = value_str.replace(',', '')
    
    # Handle European format where comma is decimal separator
    # If there are multiple dots, assume comma is decimal separator
    if value_str.count('.') > 1:
        # Keep only the last dot as decimal separator
        parts = value_str.split('.')
        value_str = ''.join(parts[:-1]) + '.' + parts[-1]
    
    # Try to parse as decimal
    try:
        return Decimal(value_str)
    except (InvalidOperation, ValueError):
        # If still fails, try to remove any remaining non-numeric characters except . and -
        import re
        cleaned = re.sub(r'[^\d.-]', '', value_str)
        if cleaned:
            try:
                return Decimal(cleaned)
            except (InvalidOperation, ValueError):
                return None
        return None

def parse_period_header(period_header):
    """Parse period headers in various formats to date objects."""
    if not period_header:
        return None
    
    # If it's already a datetime object, convert it directly
    if isinstance(period_header, datetime):
        # Validate year range
        if 2020 <= period_header.year <= 2099:
            # Always use first day of month
            return period_header.replace(day=1).date()
        return None
    
    # If it's already a date object, validate and return
    if isinstance(period_header, date):
        if 2020 <= period_header.year <= 2099:
            return period_header.replace(day=1)
        return None
    
    # Handle string formats
    period_str = str(period_header).strip()
    
    # Try different date formats
    formats_to_try = [
        '%Y-%m', '%m/%Y', '%Y/%m', '%b-%y', '%y-%b', '%B %Y', '%Y %B'
    ]
    
    for fmt in formats_to_try:
        try:
            parsed_date = datetime.strptime(period_str, fmt)
            # If year is 2 digits, assume 20xx
            if parsed_date.year < 100:
                parsed_date = parsed_date.replace(year=2000 + parsed_date.year)
            # Validate year range
            if 2020 <= parsed_date.year <= 2099:
                # Always use first day of month
                return parsed_date.replace(day=1).date()
        except ValueError:
            continue
    
    return None

def convert_month_year_to_date_range(month, year):
    """Convert separate month and year to date range (first day to last day of month)."""
    if not month or not year:
        return None, None
    
    try:
        month_int = int(month)
        year_int = int(year)
        first_day = date(year_int, month_int, 1)
        last_day = date(year_int, month_int, calendar.monthrange(year_int, month_int)[1])
        return first_day, last_day
    except (ValueError, AttributeError):
        return None, None

@login_required
def home(request):
    """Home page view with navigation and active states for the USA map."""
    active_state_qs = ActiveState.objects.filter(is_active=True).order_by('state_name')
    active_state_codes = list(active_state_qs.values_list('state_code', flat=True))
    total_deal_count = sum(state.deal_count for state in active_state_qs)
    total_deal_volume = sum(state.deal_volume for state in active_state_qs)

    return render(request, 'core/home.html', {
        'active_states': active_state_codes,
        'state_deal_rows': active_state_qs,
        'state_deal_totals': {
            'deal_count': total_deal_count,
            'deal_volume': total_deal_volume,
        },
    })


@login_required
def manage_active_states(request):
    """Simple UI for maintaining ActiveState rows used by the home page map."""

    states = list(ActiveState.objects.order_by('state_name'))
    totals = {
        'deal_count': sum(state.deal_count for state in states),
        'deal_volume': sum(state.deal_volume for state in states),
    }

    edit_state = None
    form = None

    if request.method == 'POST':
        if 'delete_id' in request.POST:
            state_to_delete = get_object_or_404(ActiveState, pk=request.POST['delete_id'])
            state_to_delete.delete()
            messages.success(request, f"Removed {state_to_delete.state_code} from Active States.")
            return redirect('core:manage_active_states')

        state_id = request.POST.get('state_id')
        instance = None
        if state_id:
            instance = get_object_or_404(ActiveState, pk=state_id)
            edit_state = instance

        form = ActiveStateForm(request.POST, instance=instance)
        if form.is_valid():
            saved_state = form.save()
            action = 'Updated' if state_id else 'Added'
            messages.success(request, f"{action} {saved_state.state_code} successfully.")
            return redirect('core:manage_active_states')
    else:
        edit_id = request.GET.get('edit')
        if edit_id:
            edit_state = get_object_or_404(ActiveState, pk=edit_id)
            form = ActiveStateForm(instance=edit_state)

    if form is None:
        form = ActiveStateForm()

    context = {
        'form': form,
        'states': states,
        'totals': totals,
        'editing_state': edit_state,
    }
    return render(request, 'core/manage_active_states.html', context)

@login_required
def chart_of_accounts_view(request):
    """View for displaying Chart of Accounts with search and hierarchical display."""
    search_query = request.GET.get('search', '')
    accounts = ChartOfAccounts.objects.all().order_by('sort_order')
    
    if search_query:
        accounts = accounts.filter(
            Q(account_code__icontains=search_query) |
            Q(account_name__icontains=search_query) |
            Q(account_type__icontains=search_query) |
            Q(parent_category__icontains=search_query) |
            Q(sub_category__icontains=search_query)
        )
    
    hierarchical_accounts = []
    parent_categories = {}
    
    for account in accounts:
        if account.parent_category:
            if account.parent_category not in parent_categories:
                parent_categories[account.parent_category] = []
            parent_categories[account.parent_category].append(account)
        else:
            hierarchical_accounts.append(account)
    
    for parent_category, sub_accounts in parent_categories.items():
        parent_header = next((acc for acc in hierarchical_accounts if acc.account_name == parent_category), None)
        if not parent_header:
            hierarchical_accounts.append(ChartOfAccounts(
                account_name=parent_category,
                account_type='',
                parent_category='',
                sub_category='',
                sort_order=0,
                is_header=True
            ))
        hierarchical_accounts.extend(sub_accounts)
    
    context = {
        'accounts': hierarchical_accounts,
        'search_query': search_query
    }
    return render(request, 'core/chart_of_accounts_simple.html', context)

@csrf_exempt
@login_required
def download_chart_of_accounts(request):
    """Download Chart of Accounts as CSV/Excel."""
    accounts = ChartOfAccounts.objects.all().order_by('sort_order')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="chart_of_accounts.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Sort Order', 'Account Code', 'Account Name', 'Type', 'Parent Category', 'Sub Category'])
    
    for account in accounts:
        writer.writerow([
            account.sort_order,
            account.account_code or '',
            account.account_name,
            account.account_type or '',
            account.parent_category or '',
            account.sub_category or ''
        ])
    
    return response

@login_required
def upload_chart_of_accounts(request):
    """Upload Chart of Accounts from CSV/Excel file."""
    if request.method == 'POST':
        uploaded_file = request.FILES['file']
        replace_existing = request.POST.get('replace_existing') == 'on'

        try:
            if replace_existing:
                existing_count = ChartOfAccounts.objects.count()
                if existing_count > 0:
                    ChartOfAccounts.objects.all().delete()
                    logger.info(f"Deleted {existing_count} existing ChartOfAccounts records for replacement")
                    messages.warning(request, f'Deleted {existing_count} existing Chart of Accounts records. Proceeding with import.')
                else:
                    messages.info(request, 'No existing Chart of Accounts to replace.')

            # Read file
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            
            # Check column count
            if len(df.columns) < 6:
                messages.error(request, f'File must have at least 6 columns. Found {len(df.columns)} columns.')
                return render(request, 'core/upload_chart_of_accounts.html')
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Skip completely empty rows
                    if row.isna().all():
                        continue
                        
                    sort_order = int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0
                    account_code = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
                    account_name = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
                    account_type = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                    parent_category = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
                    sub_category = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
                    
                    # Skip rows without account name
                    if not account_name:
                        continue

                    if not replace_existing and account_code and ChartOfAccounts.objects.filter(account_code=account_code).exists():
                        errors.append(f"Row {index + 2}: Account Code '{account_code}' already exists")
                        error_count += 1
                        continue

                    # Determine if this is a header row (no account code or specific account types)
                    is_header = not account_code or account_type in ['', 'HEADER', 'TOTAL']
                    
                    ChartOfAccounts.objects.create(
                        sort_order=sort_order,
                        account_code=account_code if account_code else None,
                        account_name=account_name,
                        account_type=account_type if account_type else '',
                        parent_category=parent_category,
                        sub_category=sub_category,
                        formula='',
                        is_header=is_header
                    )
                    success_count += 1
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1
            
            if success_count > 0:
                if replace_existing:
                    messages.success(request, f'Successfully replaced Chart of Accounts with {success_count} new records.')
                else:
                    messages.success(request, f'Successfully added/updated {success_count} records.')
            
            if error_count > 0:
                error_message = f'Encountered {error_count} errors during import. '
                if errors:
                    error_message += 'First few errors: ' + '; '.join(errors[:5])
                messages.warning(request, error_message)
                
        except Exception as e:
            logger.error(f'Error processing Chart of Accounts file: {str(e)}')
            messages.error(request, f'Error processing file: {str(e)}')
    
    return render(request, 'core/upload_chart_of_accounts.html')

@login_required
def upload_financial_data(request):
    """Upload Financial Data from CSV/Excel file."""
    companies = Company.objects.all().order_by('name')
    
    if request.method == 'POST':
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        try:
            uploaded_file = request.FILES['file']
            company_id = request.POST.get('company')
            data_type_raw = request.POST.get('data_type', 'actual')
            data_type = data_type_raw.capitalize() if data_type_raw else 'Actual'
            
            if not company_id:
                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': 'Please select a company.'})
                else:
                    messages.error(request, 'Please select a company.')
                    return render(request, 'core/upload_financial_data.html', {'companies': companies})
            
            company = Company.objects.get(id=company_id)
            
            # Read file
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            # Check minimum columns
            if len(df.columns) < 2:
                error_msg = 'File must have at least 2 columns: Account Code and at least one period column.'
                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': error_msg})
                else:
                    messages.error(request, error_msg)
                    return render(request, 'core/upload_financial_data.html', {'companies': companies})
            
            # Parse period columns (starting from column 2)
            period_columns = []
            debug_info = []
            for col in df.columns[1:]:
                debug_info.append(f"Column: '{col}' (type: {type(col)})")
                period_date = parse_period_header(col)
                if period_date:
                    period_columns.append((col, period_date))
                    debug_info.append(f"  -> Parsed as: {period_date}")
                else:
                    debug_info.append(f"  -> Failed to parse")
            
            if not period_columns:
                error_msg = f'No valid period columns found. Please ensure column headers are in format like "Jan-24", "2024-01", etc. Debug info: {" | ".join(debug_info)}'
                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': error_msg})
                else:
                    messages.error(request, error_msg)
                    return render(request, 'core/upload_financial_data.html', {'companies': companies})
            
            # Check for existing data
            existing_periods = []
            for col, period_date in period_columns:
                existing_data = FinancialData.objects.filter(
                    company=company,
                    period=period_date,
                    data_type=data_type
                )
                if existing_data.exists():
                    existing_periods.append(col)
            
            # If there's existing data and no confirmation, ask for confirmation
            if existing_periods and not request.POST.get('confirm_overwrite'):
                if is_ajax:
                    return JsonResponse({
                        'status': 'confirmation_needed',
                        'message': f'Data already exists for periods: {", ".join(existing_periods)}. Do you want to overwrite it?'
                    })
                else:
                    messages.warning(request, f'Data already exists for periods: {", ".join(existing_periods)}. Please confirm overwrite.')
                    return render(request, 'core/upload_financial_data.html', {'companies': companies})
            
            # Collect account codes from the uploaded file
            uploaded_account_codes = set()
            for index, row in df.iterrows():
                account_code_raw = row.iloc[0]
                if pd.notna(account_code_raw):
                    if isinstance(account_code_raw, float):
                        account_code = str(int(account_code_raw))
                    else:
                        account_code = str(account_code_raw).strip()
                    if account_code:
                        uploaded_account_codes.add(account_code)
            
            # Create backup before overwriting if there's existing data
            if existing_periods:
                backup_data = []
                for col, period_date in period_columns:
                    if col in existing_periods:
                        existing_records = FinancialData.objects.filter(
                            company=company,
                            period=period_date,
                            data_type=data_type,
                            account_code__in=uploaded_account_codes  # Backup только тех записей, которые будут удалены
                        )
                        for record in existing_records:
                            backup_data.append({
                                'account_code': record.account_code,
                                'amount': float(record.amount),
                                'period': record.period.isoformat()
                            })
                
                if backup_data:
                    DataBackup.objects.create(
                        company=company,
                        data_type=data_type,
                        periods=json.dumps([col for col, _ in period_columns if col in existing_periods]),
                        backup_data=backup_data,
                        user=request.user.username if request.user.is_authenticated else 'Anonymous',
                        description=f"Backup before upload on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                    backup_msg = f'Backup created for existing data in periods: {", ".join(existing_periods)}. You can restore previous data from Admin panel.'
                    if is_ajax:
                        # We'll include this in the success message
                        pass
                    else:
                        messages.warning(request, backup_msg)
                
                # Delete ONLY data for account codes that are in the uploaded file
                for col, period_date in period_columns:
                    if col in existing_periods:
                        FinancialData.objects.filter(
                            company=company,
                            period=period_date,
                            data_type=data_type,
                            account_code__in=uploaded_account_codes  # ВАЖНО: удаляем только эти коды
                        ).delete()
            
            success_count = 0
            error_count = 0
            errors = []
            debug_info = []
            
            for index, row in df.iterrows():
                try:
                    account_code_raw = row.iloc[0]
                    if pd.notna(account_code_raw):
                        # Remove .0 suffix if it's a float, then convert to string
                        if isinstance(account_code_raw, float):
                            account_code = str(int(account_code_raw))
                        else:
                            account_code = str(account_code_raw).strip()
                    else:
                        account_code = ''
                    
                    debug_info.append(f"Row {index + 2}: Account code = '{account_code}' (raw: {account_code_raw})")
                    
                    if not account_code:
                        debug_info.append(f"  -> Skipping: empty account code")
                        continue
                    
                    # Verify account exists in ChartOfAccounts (ONLY by account_code)
                    try:
                        chart_account = ChartOfAccounts.objects.get(account_code=account_code)
                        debug_info.append(f"  -> Found in Chart of Accounts: {chart_account.account_name}")
                    except ChartOfAccounts.DoesNotExist:
                        errors.append(f"Row {index + 2}: Account code '{account_code}' not found in Chart of Accounts")
                        error_count += 1
                        debug_info.append(f"  -> ERROR: Account not found")
                        continue
                    
                    # Process each period column
                    for col, period_date in period_columns:
                        amount_value = row[col]
                        debug_info.append(f"  -> Period {period_date}: value = {amount_value} (type: {type(amount_value)})")
                        
                        if pd.notna(amount_value):
                            cleaned_amount = clean_number_value(amount_value)
                            debug_info.append(f"    -> Cleaned amount: {cleaned_amount}")
                            
                            if cleaned_amount is not None:
                                FinancialData.objects.create(
                                    company=company,
                                    account_code=account_code,
                                    period=period_date,
                                    amount=cleaned_amount,
                                    data_type=data_type
                                )
                                success_count += 1
                                debug_info.append(f"    -> SUCCESS: Created record")
                            else:
                                debug_info.append(f"    -> FAILED: clean_number_value returned None")
                        else:
                            debug_info.append(f"    -> SKIPPED: pd.notna returned False")
                
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1
                    debug_info.append(f"  -> EXCEPTION: {str(e)}")
            
            # Prepare response message
            if success_count > 0:
                success_msg = f'Successfully uploaded {success_count} records for {len(period_columns)} periods.'
                if existing_periods:
                    success_msg += f' Backup created for overwritten data.'
                if errors:
                    success_msg += f' Encountered {len(errors)} errors. Please check the data format.'
                
                if is_ajax:
                    return JsonResponse({'status': 'success', 'message': success_msg})
                else:
                    messages.success(request, success_msg)
            else:
                error_msg = f'No valid data was uploaded. Please check your file format. Debug info: {" | ".join(debug_info[:10])}'
                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': error_msg})
                else:
                    messages.error(request, error_msg)
        
        except Exception as e:
            error_msg = f'Error processing file: {str(e)}'
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': error_msg})
            else:
                messages.error(request, error_msg)
    
    return render(request, 'core/upload_financial_data.html', {'companies': companies})

@csrf_exempt
@login_required
def download_financial_data_template(request):
    """Download Financial Data template as Excel."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="financial_data_template.xlsx"'
    
    # Create sample data (removed Account Name column)
    sample_data = [
        ['4113000', 60000, 80000, 104363, 130182, 150000, 175000],
        ['5216100', 20000, 25000, 30000, 35000, 40000, 45000],
        ['6011100', 50000, 50000, 52000, 52000, 54000, 54000]
    ]
    
    # Create DataFrame (removed Account Name column)
    df = pd.DataFrame(sample_data, columns=[
        'Account Code', 'Jan-24', 'Feb-24', 'Mar-24', 'Apr-24', 'May-24', 'Jun-24'
    ])
    
    # Write to Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Financial Data', index=False)
    
    return response

@csrf_exempt
@login_required
def download_template(request):
    """Download Chart of Accounts template as CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="chart_of_accounts_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Sort Order', 'Account Code', 'Account Name', 'Type', 'Parent Category', 'Sub Category'])
    
    # Sample data
    sample_data = [
        [1, '', 'TOTAL INCOME', '', '', ''],
        [2, '4113000', 'Interest Income', 'INCOME', 'TOTAL INCOME', ''],
        [45, '', 'COST OF FUNDS', '', '', ''],
        [46, '5216100', 'Interest Expense', 'EXPENSE', 'COST OF FUNDS', ''],
        [49, 'GROSS_PROFIT', 'Gross Profit', '', '', ''],
        [60, '', 'OVERHEADS', '', '', ''],
        [95, '', 'Marketing', '', 'OVERHEADS', ''],
        [96, '6011202', 'Marketing Expense', 'EXPENSE', 'OVERHEADS', 'Marketing']
    ]
    
    for row in sample_data:
        writer.writerow(row)
    
    return response

@login_required
def pl_report_data(request):
    """P&L Report data in JSON format for AG Grid, с нормализацией месяцев и фильтром по диапазону."""
    from django.conf import settings

    # Get feature flag status
    salary_module_enabled = getattr(settings, 'ENABLE_SALARY_MODULE', False)
    from_month = request.GET.get('from_month', '')
    from_year = request.GET.get('from_year', '')
    to_month = request.GET.get('to_month', '')
    to_year = request.GET.get('to_year', '')
    data_type_raw = request.GET.get('data_type', 'actual')
    # Normalize to match database: actual->Actual, budget->Budget
    data_type = data_type_raw.capitalize() if data_type_raw else 'Actual'
    logger.info(f"Parameters: from_month={from_month}, from_year={from_year}, to_month={to_month}, to_year={to_year}, data_type={data_type}")
    # Budget map for dual-streams (period -> account_code -> amount)
    budget_values = {}

    # Помощники для нормализации месяцев
    def month_start(d: date) -> date:
        return date(d.year, d.month, 1)
    def next_month(d: date) -> date:
        if d.month == 12:
            return date(d.year + 1, 1, 1)
        return date(d.year, d.month + 1, 1)

    # Превращаем выбор пользователя в месячный диапазон [start, end_exclusive)
    from_date_start, _ = convert_month_year_to_date_range(from_month, from_year)
    _, to_date_end = convert_month_year_to_date_range(to_month, to_year)
    start = month_start(from_date_start) if from_date_start else None
    end_exclusive = next_month(to_date_end) if to_date_end else None

    # Компании
    companies = list(Company.objects.all().order_by('name'))
    # For Budget/Forecast show ALL companies including budget-only, for Actual exclude budget-only
    if data_type and data_type.lower() in ['budget', 'forecast']:
        display_companies = [c for c in companies if not getattr(c, 'is_budget_only', False)]
    else:
        display_companies = [c for c in companies if not getattr(c, 'is_budget_only', False)]
    display_company_codes = {c.code for c in display_companies}
    logger.info(f"Found {len(companies)} companies.")

    # ВАЖНОЕ ИЗМЕНЕНИЕ: Фильтруем только P&L счета (INCOME и EXPENSE)
    chart_accounts_all = list(ChartOfAccounts.objects.filter(
        account_type__in=['INCOME', 'EXPENSE']
    ).order_by('sort_order'))
    chart_accounts = [a for a in chart_accounts_all if (a.account_code or '').strip()]
    logger.info(f"P&L ChartOfAccounts: total={len(chart_accounts_all)}, with_code={len(chart_accounts)}")
    
    # Получаем список всех P&L account_codes для фильтрации
    pl_account_codes = [a.account_code for a in chart_accounts if a.account_code]

    # Периоды: берём только там, где реально есть P&L данные
    try:
        if is_enabled('PL_BUDGET_PARALLEL'):
            # Define company sets for dual streams
            actual_companies = [c for c in companies if not getattr(c, 'is_budget_only', False)]
            budget_company = next((c for c in companies if getattr(c, 'is_budget_only', False)), None)

            # Actual periods (always from 'actual' stream)
            ap_q = FinancialData.objects.filter(
                company__in=actual_companies,
                data_type='Actual',
                account_code__in=pl_account_codes,
            )
            if start:
                ap_q = ap_q.filter(period__gte=start)
            if end_exclusive:
                ap_q = ap_q.filter(period__lt=end_exclusive)
            actual_periods = ap_q.values_list('period', flat=True).distinct()

            # Budget/Forecast periods (from current filter)
            if budget_company:
                bp_q = FinancialData.objects.filter(
                    company=budget_company,
                    data_type=data_type,
                    account_code__in=pl_account_codes,
                )
                if start:
                    bp_q = bp_q.filter(period__gte=start)
                if end_exclusive:
                    bp_q = bp_q.filter(period__lt=end_exclusive)
                budget_periods = bp_q.values_list('period', flat=True).distinct()
            else:
                budget_periods = []

            periods = sorted(list(set(actual_periods) | set(budget_periods)))
            logger.info(f"Dual-stream periods count: {len(periods)} (actual+budget)")
        else:
            q = FinancialData.objects.filter(
                data_type=data_type,
                account_code__in=pl_account_codes  # Фильтруем только P&L счета
            )
            if start:
                q = q.filter(period__gte=start)
            if end_exclusive:
                q = q.filter(period__lt=end_exclusive)
            if companies:
                q = q.filter(company_id__in=[c.id for c in companies])
            periods = list(q.values_list('period', flat=True).distinct().order_by('period'))
            logger.info(f"Found {len(periods)} periods with P&L data.")

            # Если периодов нет, проверяем есть ли вообще P&L данные
            if not periods:
                all_pl_periods = list(FinancialData.objects.filter(
                    account_code__in=pl_account_codes
                ).values_list('period', flat=True).distinct().order_by('period'))
                logger.warning(f"No P&L data in selected range. Available P&L periods: {all_pl_periods[:10]}")

                # Предлагаем использовать доступный диапазон
                if all_pl_periods:
                    suggested_start = all_pl_periods[0].strftime('%B %Y')
                    suggested_end = all_pl_periods[-1].strftime('%B %Y')
                    return JsonResponse({
                        'columnDefs': [],
                        'rowData': [],
                        'error': f'No P&L data found for selected period. P&L data is available from {suggested_start} to {suggested_end}',
                        'available_range': {
                            'start': all_pl_periods[0].strftime('%Y-%m-%d'),
                            'end': all_pl_periods[-1].strftime('%Y-%m-%d')
                        }
                    })
    except Exception as e:
        logger.error(f"Error fetching periods: {e}")
        periods = []

    if not periods:
        logger.warning("No P&L periods found, returning empty data.")
        return JsonResponse({
            'columnDefs': [],
            'rowData': [],
            'error': 'No P&L data found. Please check if Income and Expense accounts are properly loaded.'
        })

    # Загружаем только P&L данные за выбранные периоды и компании
    if is_enabled('PL_BUDGET_PARALLEL'):
        # Actual companies (non-budget-only) for actual stream
        actual_companies = [c for c in companies if not getattr(c, 'is_budget_only', False)]
        budget_company = next((c for c in companies if getattr(c, 'is_budget_only', False)), None)

        actual_financial_data = list(
            FinancialData.objects.filter(
                data_type='Actual',
                period__in=periods,
                company__in=actual_companies,
                account_code__in=pl_account_codes
            ).select_related('company')
        )

        budget_financial_data = []
        if budget_company:
            budget_financial_data = list(
                FinancialData.objects.filter(
                    data_type=data_type,  # respect current filter (budget/forecast)
                    period__in=periods,
                    company=budget_company,
                    account_code__in=pl_account_codes
                ).select_related('company')
            )

        all_financial_data = actual_financial_data + budget_financial_data
        logger.info(f"Dual-stream loaded records: actual={len(actual_financial_data)}, budget={len(budget_financial_data)}")
    else:
        all_financial_data = list(
            FinancialData.objects.filter(
                data_type=data_type,
                period__in=periods,
                company_id__in=[c.id for c in companies],
                account_code__in=pl_account_codes  # Только P&L счета
            ).select_related('company')
        )
    logger.info(f"Found {len(all_financial_data)} P&L financial data records.")
    
    # Добавляем детальное отладочное логирование
    if all_financial_data:
        sample = all_financial_data[0]
        logger.info(f"Sample record: company={sample.company.code} (id={sample.company.id}), account={sample.account_code}, amount={sample.amount}, period={sample.period}")
    
    # Проверяем данные по компаниям
    for c in companies:
        company_data_count = len([fd for fd in all_financial_data if fd.company.id == c.id])
        logger.info(f"Company {c.code} (id={c.id}): {company_data_count} records")
    
    # Получаем список компаний которые реально имеют данные
    companies_with_data = list(set(fd.company for fd in all_financial_data))
    if not companies_with_data:
        logger.warning("No companies with data found, using all companies as fallback")
        companies_with_data = companies
    else:
        logger.info(f"Companies with data: {[c.code for c in companies_with_data]}")

    # Parallel logic: choose P&L companies for calculations (exclude budget-only)
    pl_companies = companies_with_data
    if is_enabled('PL_BUDGET_PARALLEL'):
        pl_companies = [c for c in companies if not getattr(c, 'is_budget_only', False)]
        logger.info(f"PL companies for P&L calculations: {[c.code for c in pl_companies]}")
        # Also include budget-only company for structure initialization
        budget_company = next((c for c in companies if getattr(c, 'is_budget_only', False)), None)
        all_report_companies = pl_companies + ([budget_company] if budget_company else [])

    # Индексация: period -> company_code -> account_code -> amount
    financial_data = {}
    # Define a soft, readable palette (no reds) and map companies deterministically
    palette = ['#E6F3FF', '#E8F5E9', '#F0F4FF', '#E6F7F7', '#F6F8E7', '#F0E6FF', '#F5F5F5']
    color_by_company = {c.id: palette[i % len(palette)] for i, c in enumerate(display_companies)}

    for p in periods:
        financial_data[p] = {}
        if is_enabled('PL_BUDGET_PARALLEL'):
            # Initialize keys for both actual companies and budget-only company
            for c in all_report_companies:
                financial_data[p][c.code] = {}
        else:
            for c in pl_companies:  # Используем компании для P&L расчета
                financial_data[p][c.code] = {}
    
    # Заполняем financial_data с проверками (dual streams under feature flag)
    if is_enabled('PL_BUDGET_PARALLEL'):
        # Populate actual stream into financial_data
        for fd in actual_financial_data:
            p = fd.period
            ccode = fd.company.code
            logger.debug(f"Actual stream: period={p}, company={ccode}, account={fd.account_code}, amount={fd.amount}")
            # Fallback protection for missing keys
            if p not in financial_data:
                financial_data[p] = {}
            if ccode not in financial_data[p]:
                financial_data[p][ccode] = {}
            financial_data[p][ccode][fd.account_code] = fd.amount

        # Build budget-only mapping per period/account (do not mix into financial_data)
        for fd in budget_financial_data:
            p = fd.period
            acc = fd.account_code
            budget_values.setdefault(p, {})
            # Sum if multiple entries per period/account
            prev = budget_values[p].get(acc, 0)
            budget_values[p][acc] = prev + (fd.amount or 0)
            # print(f"DEBUG: Added to budget_values - period: {p}, account: {acc}, amount: {fd.amount}")  # debug
        
        # Debug: Show what's in budget_values
        # print(f"DEBUG BUDGET_VALUES: Total periods with budget data: {len(budget_values)}")
        for period, accounts in budget_values.items():
            # print(f"DEBUG BUDGET_VALUES: Period {period}: {len(accounts)} accounts")
            for acc, amount in accounts.items():
                # print(f"DEBUG BUDGET_VALUES:   {acc}: {amount}")
                pass  # Ensure loop has a body to avoid IndentationError
    else:
        for fd in all_financial_data:
            p = fd.period
            ccode = fd.company.code
            logger.debug(f"Processing: period={p}, company={ccode}, account={fd.account_code}, amount={fd.amount}")
            if p in financial_data and ccode in financial_data[p]:
                financial_data[p][ccode][fd.account_code] = fd.amount
            else:
                logger.warning(f"Failed to add: period={p}, company={ccode} not found in financial_data structure")
    
    # Track which company-period columns actually carry data
    non_zero_company_periods = set()
    for period, company_map in financial_data.items():
        period_key = period.strftime('%Y-%m')
        for company_code, accounts_map in company_map.items():
            if company_code not in display_company_codes:
                continue
            if any(amount for amount in accounts_map.values()):
                non_zero_company_periods.add((period_key, company_code))


    # Группировка COA по sub_category для структуры
    grouped_data = {}
    for acc in chart_accounts_all:
        sub_category = acc.sub_category or 'UNCATEGORIZED'
        grouped_data.setdefault(sub_category, []).append(acc)

    # Get P&L subcategories ordered by sort_order from database
    pl_subcategories = ChartOfAccounts.objects.filter(
        account_type__in=['INCOME', 'EXPENSE'],
        sub_category__isnull=False
    ).values('sub_category').annotate(
        min_sort_order=Min('sort_order')
    ).order_by('min_sort_order')

    # Build the ordered list of subcategories
    correct_order = [item['sub_category'] for item in pl_subcategories]
    all_sub = list(grouped_data.keys())
    pl_structure = [c for c in correct_order if c in all_sub] + [c for c in all_sub if c not in correct_order]
    logger.info(f"Ordered P&L sub categories: {pl_structure}")

    # Сборка отчета
    def build_style_token(source_text, fallback=''):
        base = source_text or fallback or ''
        return slugify(base) if base else ''

    report_data = []
    debug_info = {
        'periods_count': len(periods),
        'companies_count': len(companies),
        'pl_accounts_count': len(chart_accounts),
        'financial_data_count': len(all_financial_data),
        'periods': [p.strftime('%Y-%m-%d') for p in periods[:6]],
        'companies': [c.code for c in companies],
        'companies_with_data': [c.code for c in companies_with_data],
        'companies_without_data': [c.code for c in companies if c not in companies_with_data]
    }
    
    # Добавляем счетчики по типам
    income_count = ChartOfAccounts.objects.filter(account_type='INCOME', account_code__in=pl_account_codes).count()
    expense_count = ChartOfAccounts.objects.filter(account_type='EXPENSE', account_code__in=pl_account_codes).count()
    debug_info['income_accounts'] = income_count
    debug_info['expense_accounts'] = expense_count
    
    if all_financial_data:
        debug_info['sample_financial_data'] = [{
            'company': all_financial_data[0].company.code,
            'account': all_financial_data[0].account_code,
            'period': str(all_financial_data[0].period),
            'amount': float(all_financial_data[0].amount)
        }]

    # (Removed visual REVENUE section header to simplify layout)
    
    # Обрабатываем Income счета
    income_accounts = [a for a in chart_accounts if a.account_type == 'INCOME']
    total_revenue_snapshot = None
    gross_profit_inserted = False
    for sub_category in pl_structure:
        if sub_category not in grouped_data:
            continue
            
        # Проверяем есть ли Income счета в этой категории
        category_accounts = [a for a in grouped_data[sub_category] if a.account_type == 'INCOME' and a.account_code]
        if not category_accounts:
            continue

        normalized_sub_category = (sub_category or '').strip().upper()

        # Подзаголовок
        # Get sort_order for this subcategory
        subcategory_sort_order = ChartOfAccounts.objects.filter(
            sub_category=sub_category,
            account_type='INCOME'
        ).aggregate(min_sort=Min('sort_order'))['min_sort'] or 0
        style_token = build_style_token(sub_category)
        
        report_data.append({
            'type': 'sub_header',
            'account_name': sub_category,
            'account_code': '',
            'periods': {},
            'grand_totals': {},
            'section': 'income',
            'sort_order': subcategory_sort_order,
            'level': 1,
            'styleToken': style_token
        })

        # Счета
        for acc in category_accounts:
            row = {
                'type': 'account',
                'account_name': acc.account_name,
                'account_code': acc.account_code,
                'periods': {},
                'grand_totals': {},
                'section': 'income',
                'sort_order': acc.sort_order,
                'level': 2,
                'styleToken': style_token
            }
            has_non_zero_value = False
            # Помесячно
            for p in periods:
                row['periods'][p] = {}
                period_total = Decimal('0')
                for c in companies_with_data:  # Используем только компании с данными
                    amount = financial_data[p][c.code].get(acc.account_code, Decimal('0'))
                    row['periods'][p][c.code] = float(amount or Decimal('0'))
                    period_total += amount or Decimal('0')
                    if amount != 0:
                        has_non_zero_value = True
                row['periods'][p]['TOTAL'] = float(period_total or Decimal('0'))

            # Гранд тоталы
            for c in companies_with_data:  # Используем только компании с данными
                grand_total = sum(financial_data[p][c.code].get(acc.account_code, Decimal('0')) for p in periods)
                row['grand_totals'][c.code] = float(grand_total or Decimal('0'))
            overall = sum(
                sum(financial_data[p][c.code].get(acc.account_code, Decimal('0')) for c in companies_with_data)
                for p in periods
            )
            row['grand_totals']['TOTAL'] = float(overall or Decimal('0'))

            if has_non_zero_value:
                report_data.append(row)

        # Субитог секции
        sub_total = {
            'type': 'sub_total',
            'account_name': f'Total {sub_category}',
            'account_code': '',
            'periods': {},
            'grand_totals': {},
            'section': 'income',
            'sort_order': subcategory_sort_order,
            'level': 1,
            'styleToken': style_token
        }
        for p in periods:
            sub_total['periods'][p] = {}
            period_total = Decimal('0')
            for c in companies_with_data:  # Используем только компании с данными
                company_total = sum(
                    financial_data[p][c.code].get(a.account_code, Decimal('0'))
                    for a in category_accounts
                )
                sub_total['periods'][p][c.code] = float(company_total or Decimal('0'))
                period_total += company_total or Decimal('0')
            sub_total['periods'][p]['TOTAL'] = float(period_total or Decimal('0'))
            # Budget subtotal per period (Budget/Forecast only)
            if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
                budget_subtotal = Decimal('0')
                for acc in category_accounts:
                    raw_budget = budget_values.get(p, {}).get(acc.account_code)
                    if raw_budget is not None:
                        try:
                            budget_subtotal += (raw_budget if isinstance(raw_budget, Decimal) else Decimal(str(raw_budget)))
                        except Exception:
                            pass
                if budget_subtotal != 0:
                    try:
                        sub_total['periods'][p]['Budget'] = float(budget_subtotal)
                    except Exception:
                        sub_total['periods'][p]['Budget'] = None

        for c in companies_with_data:  # Используем только компании с данными
            gtot = sum(
                sum(financial_data[p][c.code].get(a.account_code, Decimal('0')) for a in category_accounts)
                for p in periods
            )
            sub_total['grand_totals'][c.code] = float(gtot or Decimal('0'))
        overall = sum(
            sum(sum(financial_data[p][c.code].get(a.account_code, Decimal('0')) for a in category_accounts) for c in companies_with_data)
            for p in periods
        )
        sub_total['grand_totals']['TOTAL'] = float(overall or Decimal('0'))

        report_data.append(sub_total)

    # Total Revenue
    total_revenue_row = {
        'type': 'total',
        'account_name': 'TOTAL REVENUE',
        'account_code': '',
        'periods': {},
        'grand_totals': {},
        'section': 'income',
        'level': 0,
        'sort_order': 0,
        'styleToken': build_style_token('TOTAL REVENUE')
    }
    for p in periods:
        total_revenue_row['periods'][p] = {}
        period_total = Decimal('0')
        for c in pl_companies:  # Используем только компании с данными
            company_total = sum(
                financial_data[p][c.code].get(a.account_code, Decimal('0'))
                for a in income_accounts
            )
            total_revenue_row['periods'][p][c.code] = float(company_total or Decimal('0'))
            period_total += company_total or Decimal('0')
        total_revenue_row['periods'][p]['TOTAL'] = float(period_total or Decimal('0'))
        # Budget total revenue by period (Budget/Forecast only)
        if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
            budget_total = Decimal('0')
            # print(f"DEBUG TOTAL REVENUE: Calculating Budget for period {p}")
            for a in income_accounts:
                raw_budget = budget_values.get(p, {}).get(a.account_code)
                if raw_budget is not None:
                    try:
                        budget_total += (raw_budget if isinstance(raw_budget, Decimal) else Decimal(str(raw_budget)))
                        # print(f"DEBUG TOTAL REVENUE: Added {a.account_code} = {raw_budget}, running total = {budget_total}")
                    except Exception:
                        pass
            # print(f"DEBUG TOTAL REVENUE: Final budget_total for period {p} = {budget_total}")
            if budget_total != 0:
                try:
                    total_revenue_row['periods'][p]['Budget'] = float(budget_total)
                    # print(f"DEBUG TOTAL REVENUE: Set total_revenue_row['periods'][{p}]['Budget'] = {float(budget_total)}")
                except Exception:
                    total_revenue_row['periods'][p]['Budget'] = None
                    # print(f"DEBUG TOTAL REVENUE: Set total_revenue_row['periods'][{p}]['Budget'] = None (exception)")
            else:
                # print(f"DEBUG TOTAL REVENUE: budget_total is 0, not setting Budget value")
                pass
    # Grand totals for revenue
    for c in pl_companies:  # Используем только компании с данными
        gtot = sum(
            sum(financial_data[p][c.code].get(a.account_code, Decimal('0')) for a in income_accounts)
            for p in periods
        )
        total_revenue_row['grand_totals'][c.code] = float(gtot or Decimal('0'))
    overall_revenue = sum(
        sum(sum(financial_data[p][c.code].get(a.account_code, Decimal('0')) for a in income_accounts) for c in pl_companies)
        for p in periods
    )
    total_revenue_row['grand_totals']['TOTAL'] = float(overall_revenue or Decimal('0'))
    total_revenue_snapshot = copy.deepcopy(total_revenue_row)

    report_data.append(total_revenue_row)
    
    # Добавляем раздел EXPENSES
    expense_total_row = {
        'type': 'section_header',
        'account_name': 'EXPENSES',
        'account_code': '',
        'periods': {},
        'grand_totals': {},
        'section': 'expense',
        'level': 0,
        'sort_order': 0,
        'styleToken': build_style_token('EXPENSES')
    }
    report_data.append(expense_total_row)

    # Обрабатываем Expense счета
    expense_accounts = [a for a in chart_accounts if a.account_type == 'EXPENSE']
    for sub_category in pl_structure:
        if sub_category not in grouped_data:
            continue
            
        # Проверяем есть ли Expense счета в этой категории
        category_accounts = [a for a in grouped_data[sub_category] if a.account_type == 'EXPENSE' and a.account_code]
        if not category_accounts:
            continue

        # Подзаголовок
        # Get sort_order for this subcategory
        subcategory_sort_order = ChartOfAccounts.objects.filter(
            sub_category=sub_category,
            account_type='EXPENSE'
        ).aggregate(min_sort=Min('sort_order'))['min_sort'] or 0
        style_token = build_style_token(sub_category)
        
        report_data.append({
            'type': 'sub_header',
            'account_name': sub_category,
            'account_code': '',
            'periods': {},
            'grand_totals': {},
            'section': 'expense',
            'sort_order': subcategory_sort_order,
            'level': 1,
            'styleToken': style_token
        })

        # Счета (аналогично Income)
        for acc in category_accounts:
            row = {
                'type': 'account',
                'account_name': acc.account_name,
                'account_code': acc.account_code,
                'periods': {},
                'grand_totals': {},
                'section': 'expense',
                'sort_order': acc.sort_order,
                'level': 2,
                'styleToken': style_token
            }
            has_non_zero_value = False
            # Помесячно
            for p in periods:
                row['periods'][p] = {}
                period_total = Decimal('0')
                for c in pl_companies:
                    # Diagnostic logging for key formats during lookup (expense section)
                    amount = financial_data[p][c.code].get(acc.account_code, 0)
                    row['periods'][p][c.code] = float(amount or 0)
                    period_total += amount or 0
                    if amount != 0:
                        has_non_zero_value = True
                row['periods'][p]['TOTAL'] = float(period_total or 0)

            # Гранд тоталы
            for c in pl_companies:
                grand_total = sum(financial_data[p][c.code].get(acc.account_code, 0) for p in periods)
                row['grand_totals'][c.code] = float(grand_total or 0)
            overall = sum(
                sum(financial_data[p][c.code].get(acc.account_code, 0) for c in pl_companies)
                for p in periods
            )
            row['grand_totals']['TOTAL'] = float(overall or 0)

            if has_non_zero_value:
                report_data.append(row)

        # Субитог секции
        sub_total = {
            'type': 'sub_total',
            'account_name': f'Total {sub_category}',
            'account_code': '',
            'periods': {},
            'grand_totals': {},
            'section': 'expense',
            'sort_order': subcategory_sort_order,
            'level': 1,
            'styleToken': style_token
        }
        for p in periods:
            sub_total['periods'][p] = {}
            period_total = Decimal('0')
            for c in pl_companies:
                company_total = sum(
                    financial_data[p][c.code].get(a.account_code, 0)
                    for a in category_accounts
                )
                sub_total['periods'][p][c.code] = float(company_total or 0)
                # Accumulate per-company totals into the per-period TOTAL
                period_total += company_total or 0
            # Set per-period TOTAL after summing all companies
            sub_total['periods'][p]['TOTAL'] = float(period_total or 0)
            # Budget subtotal per period (Budget/Forecast only)
            if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
                budget_subtotal = Decimal('0')
                # print(f"DEBUG SUBTOTAL: Calculating Budget for {sub_category}, period {p}")
                for acc in category_accounts:
                    raw_budget = budget_values.get(p, {}).get(acc.account_code)
                    if raw_budget is not None:
                        try:
                            budget_subtotal += (raw_budget if isinstance(raw_budget, Decimal) else Decimal(str(raw_budget)))
                            # print(f"DEBUG SUBTOTAL: Added {acc.account_code} = {raw_budget}, running total = {budget_subtotal}")
                        except Exception:
                            pass
                # print(f"DEBUG SUBTOTAL: Final budget_subtotal for {sub_category}, period {p} = {budget_subtotal}")
                if budget_subtotal != 0:
                    try:
                        sub_total['periods'][p]['Budget'] = float(budget_subtotal)
                        # print(f"DEBUG SUBTOTAL: Set sub_total['periods'][{p}]['Budget'] = {float(budget_subtotal)}")
                    except Exception:
                        sub_total['periods'][p]['Budget'] = None
                        # print(f"DEBUG SUBTOTAL: Set sub_total['periods'][{p}]['Budget'] = None (exception)")
                else:
                    # print(f"DEBUG SUBTOTAL: budget_subtotal is 0, not setting Budget value")
                    pass

        for c in pl_companies:
            gtot = sum(
                sum(financial_data[p][c.code].get(a.account_code, 0) for a in category_accounts)
                for p in periods
            )
            sub_total['grand_totals'][c.code] = float(gtot or 0)
        overall = sum(
            sum(sum(financial_data[p][c.code].get(a.account_code, 0) for a in category_accounts) for c in pl_companies)
            for p in periods
        )
        sub_total['grand_totals']['TOTAL'] = float(overall or 0)
        
        # Debug: Show what's in the subtotal row before adding to report_data
        # print(f"DEBUG SUBTOTAL ROW: Final subtotal for '{sub_total['account_name']}':")
        for p in periods:
            budget_val = sub_total['periods'][p].get('Budget')
            # print(f"DEBUG SUBTOTAL ROW:   Period {p}: Budget = {budget_val}")
        
        report_data.append(sub_total)

        normalized_sub_category = (sub_category or '').strip().upper()
        if (
            not gross_profit_inserted
            and normalized_sub_category == 'COST OF FUNDS AND FEES'
            and total_revenue_snapshot is not None
        ):
            def to_decimal(value):
                if value in (None, ''):
                    return Decimal('0')
                try:
                    return Decimal(str(value))
                except (InvalidOperation, TypeError):
                    return Decimal('0')

            gross_profit_row = {
                'type': 'total',
                'account_name': 'Gross Profit',
                'account_code': '',
                'periods': {},
                'grand_totals': {},
                'section': 'summary',
                'level': 0,
                'sort_order': (sub_total.get('sort_order') or 0) + 1,
                'styleToken': build_style_token('Gross Profit')
            }

            for p in periods:
                gross_profit_row['periods'][p] = {}
                period_total = Decimal('0')
                for c in pl_companies:
                    revenue_val = to_decimal(total_revenue_snapshot['periods'][p].get(c.code))
                    cost_val = to_decimal(sub_total['periods'][p].get(c.code))
                    gross_val = revenue_val - cost_val
                    gross_profit_row['periods'][p][c.code] = float(gross_val)
                    period_total += gross_val
                gross_profit_row['periods'][p]['TOTAL'] = float(period_total)

                if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
                    revenue_budget = to_decimal(total_revenue_snapshot['periods'][p].get('Budget'))
                    cost_budget = to_decimal(sub_total['periods'][p].get('Budget'))
                    budget_val = revenue_budget - cost_budget
                    gross_profit_row['periods'][p]['Budget'] = float(budget_val) if budget_val != 0 else None

            for c in pl_companies:
                revenue_total = to_decimal(total_revenue_snapshot['grand_totals'].get(c.code))
                cost_total = to_decimal(sub_total['grand_totals'].get(c.code))
                gross_profit_row['grand_totals'][c.code] = float(revenue_total - cost_total)

            overall_gross = (
                to_decimal(total_revenue_snapshot['grand_totals'].get('TOTAL'))
                - to_decimal(sub_total['grand_totals'].get('TOTAL'))
            )
            gross_profit_row['grand_totals']['TOTAL'] = float(overall_gross)

            if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
                revenue_budget_total = to_decimal(total_revenue_snapshot['grand_totals'].get('Budget'))
                cost_budget_total = to_decimal(sub_total['grand_totals'].get('Budget'))
                gross_budget_total = revenue_budget_total - cost_budget_total
                gross_profit_row['grand_totals']['Budget'] = (
                    float(gross_budget_total) if gross_budget_total != 0 else None
                )

            insert_index = len(report_data)
            report_data.insert(insert_index, gross_profit_row)
            gross_profit_inserted = True

    # Total Expenses
    total_expense_row = {
        'type': 'total',
        'account_name': 'TOTAL EXPENSES',
        'account_code': '',
        'periods': {},
        'grand_totals': {},
        'section': 'expense',
        'level': 0,
        'sort_order': 0,
        'styleToken': build_style_token('TOTAL EXPENSES')
    }
    for p in periods:
        total_expense_row['periods'][p] = {}
        period_total = Decimal('0')
        for c in pl_companies:
            company_total = sum(
                financial_data[p][c.code].get(a.account_code, 0)
                for a in expense_accounts
            )
            total_expense_row['periods'][p][c.code] = float(company_total or 0)
            period_total += company_total or 0
        total_expense_row['periods'][p]['TOTAL'] = float(period_total or 0)
        # Budget total expenses by period (Budget/Forecast only)
        # print(f"DEBUG TOTAL EXPENSES: Feature flag check - is_enabled('PL_BUDGET_PARALLEL')={is_enabled('PL_BUDGET_PARALLEL')}, data_type='{data_type}', data_type.lower() in ['budget', 'forecast']={data_type.lower() in ['budget', 'forecast']}")
        if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
            budget_total = Decimal('0')
            # print(f"DEBUG TOTAL EXPENSES: Calculating Budget for period {p}")
            for a in expense_accounts:
                raw_budget = budget_values.get(p, {}).get(a.account_code)
                if raw_budget is not None:
                    try:
                        budget_total += (raw_budget if isinstance(raw_budget, Decimal) else Decimal(str(raw_budget)))
                        # print(f"DEBUG TOTAL EXPENSES: Added {a.account_code} = {raw_budget}, running total = {budget_total}")
                    except Exception:
                        pass
            # print(f"DEBUG TOTAL EXPENSES: Final budget_total for period {p} = {budget_total}")
            if budget_total != 0:
                try:
                    total_expense_row['periods'][p]['Budget'] = float(budget_total)
                    # print(f"DEBUG TOTAL EXPENSES: Set total_expense_row['periods'][{p}]['Budget'] = {float(budget_total)}")
                except Exception:
                    total_expense_row['periods'][p]['Budget'] = None
                    # print(f"DEBUG TOTAL EXPENSES: Set total_expense_row['periods'][{p}]['Budget'] = None (exception)")
            else:
                # print(f"DEBUG TOTAL EXPENSES: budget_total is 0, not setting Budget value")
                pass
    # Grand totals for expenses
    for c in pl_companies:
        gtot = sum(
            sum(financial_data[p][c.code].get(a.account_code, 0) for a in expense_accounts)
            for p in periods
        )
        total_expense_row['grand_totals'][c.code] = float(gtot or 0)
    overall_expense = sum(
        sum(sum(financial_data[p][c.code].get(a.account_code, 0) for a in expense_accounts) for c in pl_companies)
        for p in periods
    )
    total_expense_row['grand_totals']['TOTAL'] = float(overall_expense or 0)
    report_data.append(total_expense_row)
    
    # NET INCOME (Revenue - Expenses)
    net_income_row = {
        'type': 'net_income',
        'account_name': 'NET INCOME',
        'account_code': '',
        'periods': {},
        'grand_totals': {},
        'section': 'summary',
        'level': 0,
        'sort_order': 0,
        'styleToken': build_style_token('NET INCOME')
    }
    for p in periods:
        net_income_row['periods'][p] = {}
        period_total = Decimal('0')
        for c in pl_companies:
            revenue = Decimal(str(total_revenue_row['periods'][p][c.code]))
            expense = Decimal(str(total_expense_row['periods'][p][c.code]))
            net = revenue - expense
            net_income_row['periods'][p][c.code] = float(net)
            period_total += net
        net_income_row['periods'][p]['TOTAL'] = float(period_total)
    # Calculate NET INCOME Budget values
    if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
        for p in periods:
            revenue_budget = total_revenue_row['periods'].get(p, {}).get('Budget', 0) or 0
            expense_budget = total_expense_row['periods'].get(p, {}).get('Budget', 0) or 0
            
            if revenue_budget or expense_budget:
                net_income_budget = revenue_budget - expense_budget
                net_income_row['periods'][p]['Budget'] = net_income_budget
                # print(f"DEBUG NET INCOME: Period {p}: Revenue {revenue_budget} - Expenses {expense_budget} = {net_income_budget}")
    # Grand totals for net income
    for c in pl_companies:
        revenue = Decimal(str(total_revenue_row['grand_totals'][c.code]))
        expense = Decimal(str(total_expense_row['grand_totals'][c.code]))
        net_income_row['grand_totals'][c.code] = float(revenue - expense)
    net_income_row['grand_totals']['TOTAL'] = float(Decimal(str(total_revenue_row['grand_totals']['TOTAL'])) - Decimal(str(total_expense_row['grand_totals']['TOTAL'])))
    report_data.append(net_income_row)

    # Колонки AG Grid
    column_defs = [
        {
            "headerName": "+",
            "field": "toggler",
            "headerComponent": "togglerHeader",
            "width": 40,
            "pinned": "left",
            "hide": False
        },
        {
            "field": "account_code",
            "colId": "account_code",
            "headerName": "A/C",
            "headerComponent": "hideableHeader",
            "width": 90,
            "pinned": "left",
            "hide": True,
            "sortable": True
        },
        {'field': 'account_name', 'headerName': 'Account Name', 'pinned': 'left', 'width': 250}
    ]
    for p in periods:
        # Build columns for this period explicitly in desired order:
        # 1) Company columns, 2) TOTAL, 3) Budget
        period_cols = []
        for c in display_companies:  # Use filtered list
            period_cols.append({
                'field': f'{p.strftime("%b-%y")}_{c.code}',
                'headerName': f'{p.strftime("%b-%y")} {c.code}',
                'width': 120,
                'type': 'numberColumnWithCommas',
                'colType': 'company',
                'periodKey': p.strftime('%Y-%m'),
                'companyCode': c.code,
                'cellStyle': {
                    'textAlign': 'right',
                    'backgroundColor': color_by_company.get(getattr(c, 'id', None), '#F5F5F5')
                }
            })
        # P&L TOTAL per period (existing)
        period_cols.append({
            'field': f'{p.strftime("%b-%y")}_TOTAL',
            'headerName': f'{p.strftime("%b-%y")} TOTAL',
            'headerComponent': 'periodToggleHeader',
            'width': 120,
            'type': 'numberColumnWithCommas',
            'colType': 'total',
            'periodKey': p.strftime('%Y-%m'),
            'cellStyle': {
                'textAlign': 'right',
                'backgroundColor': '#FFF9E6'
            }
        })
        # Add Budget/Forecast consolidated column when viewing Budget or Forecast
        if data_type and data_type.lower() in ['budget', 'forecast']:
            period_cols.append({
                'headerName': f'{p.strftime("%b-%y")} Budget',
                'field': f'{p.strftime("%b-%y")}_Budget',
                'type': 'numberColumnWithCommas',
                'cellClass': 'budget-cell',
                'colType': 'budget',
                'cellStyle': {
                    'textAlign': 'right',
                    'backgroundColor': '#F0F0FF'
                },
                'width': 120
            })
        # Extend into main column defs in the exact order
        column_defs.extend(period_cols)
        # CF Dashboard TOTAL per period removed to avoid duplicate TOTAL columns
    # Ensure visual consistency: Non-budget company grand totals first, then overall TOTAL, then budget-only grand total
    non_budget_companies = [c for c in companies if not getattr(c, 'is_budget_only', False)]
    budget_only_companies = [c for c in companies if getattr(c, 'is_budget_only', False)]

    # Grand totals for regular companies
    for c in non_budget_companies:
        column_defs.append({
            'field': f'grand_total_{c.code}',
            'headerName': f'Grand Total {c.code}',
            'colType': 'grand_company',
            'width': 120,
            'type': 'numberColumnWithCommas',
            'cellStyle': {
                'textAlign': 'right',
                'backgroundColor': color_by_company.get(getattr(c, 'id', None), '#F5F5F5')
            }
        })

    # Overall Grand Total (TOTAL) column
    column_defs.append({
        'field': 'grand_total_TOTAL',
        'headerName': 'Grand Total',
        'colType': 'grand_overall',
        'headerComponent': 'grandTotalsToggleHeader',
        'width': 120,
        'type': 'numberColumnWithCommas',
        'cellStyle': {
            'textAlign': 'right',
            'backgroundColor': '#FFF9E6'
        }
    })

    # Grand Total Budget column (only in Budget/Forecast views) — ensure added only once
    if data_type and data_type.lower() in ['budget', 'forecast']:
        if not any(col.get('field') == 'grand_total_Budget' for col in column_defs):
            column_defs.append({
                'field': 'grand_total_Budget',
                'headerName': 'Grand Total Budget',
                'colType': 'grand_budget',
                'width': 120,
                'type': 'numberColumnWithCommas',
                'cellStyle': {
                    'textAlign': 'right',
                    'backgroundColor': '#F0F0FF'
                }
            })

    # Grand total for budget-only company (if present) is skipped to avoid field name collision
    # with the overall 'grand_total_Budget' column. This per-company column is unused/empty.
    for c in budget_only_companies:
        pass

    # CF Dashboard section - loan movements and funding metrics
    cf_metrics = CFDashboardMetric.objects.filter(is_active=True).order_by('display_order')
    # Guarded date filters for CF data
    cf_query = CFDashboardData.objects.filter(company__in=companies)
    if from_date_start:
        cf_query = cf_query.filter(period__gte=from_date_start)
    if to_date_end:
        cf_query = cf_query.filter(period__lte=to_date_end)
    cf_data = cf_query.select_related('metric', 'company')
    
    # Load CF Dashboard Budget/Forecast data (consolidated, not per company)
    from .models import CFDashboardBudget
    cf_budget_query = CFDashboardBudget.objects.filter(data_type=data_type.lower())
    if from_date_start:
        cf_budget_query = cf_budget_query.filter(period__gte=from_date_start)
    if to_date_end:
        cf_budget_query = cf_budget_query.filter(period__lte=to_date_end)
    cf_budget_data = cf_budget_query
    
    # Build CF Dashboard rows
    cf_rows = []
    for metric in cf_metrics:
        row = {
            'account_code': '',
            'account_name': metric.metric_name,
            'is_cf_dashboard': True,
            'metric_id': metric.id
        }
        
        # Identify metric type
        name_val = metric.metric_name or ''
        is_cumulative_metric = 'Cumulative' in name_val
        is_ytd_metric = ('YTD' in name_val) and not is_cumulative_metric
        
        previous_total = 0  # for YTD accumulation of TOTAL only
        loans_advanced_row = None
        
        # If cumulative, try to find the already-built "Loans advanced in month" row
        if is_cumulative_metric:
            for existing_row in cf_rows:
                if 'Loans advanced in month' in (existing_row.get('account_name') or ''):
                    loans_advanced_row = existing_row
                    break
        
        # Process each period
        for period in periods:
            period_total = 0

            # Get values for each company
            for company in display_companies:  # Use filtered list
                period_key = f"{period.strftime('%b-%y')}_{company.code}"
                
                if is_cumulative_metric:
                    # For January: use input value for cumulative metric
                    if period.month == 1:
                        cf_value = cf_data.filter(
                            company=company,
                            period=period,
                            metric=metric
                        ).values_list('value', flat=True).first()
                        value = cf_value or 0
                    else:
                        # Feb-Dec: previous month cumulative + current month loans advanced
                        prev_month = period.month - 1
                        prev_period = period.replace(month=prev_month)
                        prev_key = f"{prev_period.strftime('%b-%y')}_{company.code}"
                        prev_cumulative = row.get(prev_key, 0)
                        current_loans = 0
                        if loans_advanced_row is not None:
                            current_loans = loans_advanced_row.get(period_key, 0) or 0
                        value = (prev_cumulative or 0) + (current_loans or 0)
                else:
                    # Regular or YTD metric: use input values per company-month
                    cf_value = cf_data.filter(
                        company=company,
                        period=period,
                        metric=metric
                    ).values_list('value', flat=True).first()
                    value = cf_value or 0
                
                row[period_key] = value
                if (
                    company.code in display_company_codes
                    and value not in (None, 0, 0.0)
                ):
                    non_zero_company_periods.add((period.strftime('%Y-%m'), company.code))
                period_total += value
            
            # Calculate TOTAL column
            period_total_key = f"{period.strftime('%b-%y')}_TOTAL"
            if is_ytd_metric:
                current_ytd_total = previous_total + period_total
                row[period_total_key] = current_ytd_total
                previous_total = current_ytd_total
            else:
                row[period_total_key] = period_total
            
            # Add Budget/Forecast consolidated value for this period (single column)
            if data_type and data_type.lower() in ['budget', 'forecast']:
                period_budget_key = f"{period.strftime('%b-%y')}_Budget"
                budget_value = cf_budget_data.filter(
                    metric=metric,
                    period=period
                ).values_list('value', flat=True).first()
                row[period_budget_key] = float(budget_value) if budget_value is not None else None
        
        # CF: Sum per-period Budget values into grand_total_Budget (Budget/Forecast only)
        if data_type and data_type.lower() in ['budget', 'forecast']:
            total_budget_sum = 0
            for p in periods:
                key = f"{p.strftime('%b-%y')}_Budget"
                val = row.get(key)
                if val is not None:
                    try:
                        total_budget_sum += float(val)
                    except Exception:
                        pass
            row['grand_total_Budget'] = None if total_budget_sum == 0 else float(total_budget_sum)

        cf_rows.append(row)

    # Apply visibility to company-period columns after incorporating CF data
    for col in column_defs:
        if col.get('colType') != 'company':
            continue
        period_key = col.get('periodKey')
        company_code = col.get('companyCode')
        if not period_key or not company_code:
            continue
        col['hide'] = (period_key, company_code) not in non_zero_company_periods

    active_company_codes = {company_code for _, company_code in non_zero_company_periods}
    for col in column_defs:
        if col.get('colType') != 'grand_company':
            continue
        field_name = col.get('field') or ''
        if not isinstance(field_name, str) or not field_name.startswith('grand_total_'):
            continue
        company_code = field_name.replace('grand_total_', '', 1)
        if not company_code:
            continue
        col['hide'] = company_code not in active_company_codes

    # Данные строк для AG Grid
    def make_row_key(row_dict):
        """Generate a stable identifier for a P&L row for comment mapping."""
        sort_part = row_dict.get('sort_order', 0)
        type_part = row_dict.get('type') or row_dict.get('rowType') or 'row'
        code_part = row_dict.get('account_code') or row_dict.get('styleToken') or slugify(row_dict.get('account_name', 'row'))
        return f"{type_part}__{code_part}__{sort_part}"

    row_data = []
    
    # Add CF Dashboard rows at the top
    for cf_row in cf_rows:
        cf_row.setdefault('rowType', 'cf_metric')
        cf_row.setdefault('section', 'cf_dashboard')
        cf_row.setdefault('level', 0)
        cf_row.setdefault('styleToken', build_style_token(cf_row.get('account_name')))
        cf_row.setdefault('rowKey', f"cf__{cf_row.get('styleToken') or slugify(cf_row.get('account_name', 'metric'))}")
        row_data.append(cf_row)
    
    # Add visual separator after CF Dashboard
    if cf_rows:
        row_data.append({
            'account_code': '',
            'account_name': '━' * 30 + ' P&L REPORT ' + '━' * 30,
            'is_separator': True,
            'rowType': 'separator',
            'section': 'pnl_report',
            'level': 0,
            'styleToken': build_style_token('P&L REPORT separator')
        })
    
    # Then add regular P&L rows
    for r in report_data:
        grid_row = {
            'account_code': r['account_code'],
            'account_name': r['account_name'],
            'rowType': r['type'],
            'section': r.get('section', ''),
            'sort_order': r.get('sort_order', 0),
            'level': r.get('level', 0),
            'styleToken': r.get('styleToken', '')
        }

        grid_row['rowKey'] = make_row_key(r)

        sort_order_str = str(grid_row.get('sort_order'))
        row_type = grid_row.get('rowType')
        account_name = grid_row.get('account_name')

        condition_met = (
            salary_module_enabled
            and sort_order_str == '1100'
            and row_type == 'account'
        )
        if condition_met:
            grid_row['is_salary'] = True
            grid_row['can_view_details'] = request.user.has_perm('core.view_salary_details')
        for p in periods:
            for c in non_budget_companies:
                field = f'{p.strftime("%b-%y")}_{c.code}'
                value = r['periods'].get(p, {}).get(c.code, 0)
                # Send None for zero values so grid shows empty cells
                grid_row[field] = None if value == 0 or value is None else float(value)
            field_total = f'{p.strftime("%b-%y")}_TOTAL'
            total_value = r['periods'].get(p, {}).get('TOTAL', 0)
            # Hide zeros in TOTAL columns as well
            grid_row[field_total] = None if total_value == 0 or total_value is None else float(total_value)
            # Populate consolidated Budget for P&L rows under feature flag using dual stream budget_values
            if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
                field_budget = f'{p.strftime("%b-%y")}_Budget'
                budget_amount = 0
                if r['type'] == 'account':
                    acc = r.get('account_code')
                    if acc:
                        budget_amount = float(budget_values.get(p, {}).get(acc, 0))
                elif r['type'] in ['sub_total', 'total', 'net_income']:
                    # For subtotal, total, and net_income rows, get the Budget value from the row data
                    budget_amount = r['periods'].get(p, {}).get('Budget', 0)
                    if r['type'] == 'sub_total':
                        # print(f"DEBUG GRID: Processing sub_total row '{r['account_name']}', period {p}, Budget value = {budget_amount}")
                        pass
                    elif r['type'] in ['total', 'net_income']:
                        # print(f"DEBUG GRID: Processing {r['type']} row '{r['account_name']}', period {p}, Budget value = {budget_amount}")
                        pass
                # For other row types, leave budget empty
                grid_row[field_budget] = None if not budget_amount else float(budget_amount)
                # Only show debug for subtotal and total rows, not individual accounts
                if r['type'] in ['sub_total', 'total', 'net_income']:
                    # print(f"DEBUG GRID: Set grid_row[{field_budget}] = {grid_row[field_budget]} for row type {r['type']}")
                    pass
        for c in non_budget_companies:
            field = f'grand_total_{c.code}'
            gt_val = r['grand_totals'].get(c.code, 0)
            # Hide zero company grand totals by sending None
            grid_row[field] = None if gt_val == 0 or gt_val is None else float(gt_val)
        # Overall grand total: hide zero as empty
        overall_total_value = r['grand_totals'].get('TOTAL', 0)
        grid_row['grand_total_TOTAL'] = None if overall_total_value == 0 or overall_total_value is None else float(overall_total_value)

        # P&L: Sum per-period Budget values into grand_total_Budget for all row types (Budget/Forecast only)
        if is_enabled('PL_BUDGET_PARALLEL') and data_type.lower() in ['budget', 'forecast']:
            total_budget_sum = 0
            # Only show debug for subtotal and total rows, not individual accounts
            if r['type'] in ['sub_total', 'total', 'net_income']:
                # print(f"DEBUG GRAND TOTAL: Calculating for row '{r['account_name']}' (type: {r['type']})")
                pass
            for p in periods:
                key = f"{p.strftime('%b-%y')}_Budget"
                val = grid_row.get(key)
                if val is not None:
                    try:
                        total_budget_sum += float(val)
                        if r['type'] in ['sub_total', 'total', 'net_income']:
                            # print(f"DEBUG GRAND TOTAL:   Added {key} = {val}, running total = {total_budget_sum}")
                            pass
                    except Exception:
                        pass
                else:
                    if r['type'] in ['sub_total', 'total', 'net_income']:
                        # print(f"DEBUG GRAND TOTAL:   {key} = None (skipped)")
                        pass
            grid_row['grand_total_Budget'] = None if total_budget_sum == 0 else float(total_budget_sum)
            if r['type'] in ['sub_total', 'total', 'net_income']:
                # print(f"DEBUG GRAND TOTAL: Final grand_total_Budget = {grid_row['grand_total_Budget']}")
                pass

        # Стили для разных типов строк
        if r['type'] == 'section_header':
            grid_row['cellStyle'] = {'backgroundColor': '#d4edda', 'fontWeight': 'bold', 'fontSize': '15px'}
        elif r['type'] == 'sub_header':
            grid_row['cellStyle'] = {'backgroundColor': '#f0f8ff', 'fontWeight': 'bold'}
        elif r['type'] == 'sub_total':
            grid_row['cellStyle'] = {'backgroundColor': '#fff8dc', 'fontWeight': 'bold'}
        elif r['type'] == 'total':
            grid_row['cellStyle'] = {'backgroundColor': '#e8f5e8', 'fontWeight': 'bold', 'fontSize': '14px'}
        elif r['type'] == 'net_income':
            grid_row['cellStyle'] = {'backgroundColor': '#ffeaa7', 'fontWeight': 'bold', 'fontSize': '15px'}

        row_data.append(grid_row)

        account_name_upper = (grid_row.get('account_name') or '').strip().upper()
        if account_name_upper == 'TOTAL COST OF FUNDS AND FEES':
            spacer_row = {
                'account_code': '',
                'account_name': '',
                'rowType': 'spacer',
                'section': 'spacer',
                'level': 0,
                'styleToken': 'spacer-row',
                'rowKey': f"spacer__gross_profit__{len(row_data)}",
            }
            for col in column_defs:
                field_name = col.get('field')
                if field_name:
                    spacer_row[field_name] = None
            row_data.append(spacer_row)

    # Build comment summary for visible rows/columns
    comment_summary = {}
    if row_data:
        row_keys_for_comments = {row.get('rowKey') for row in row_data if row.get('rowKey')}
        if row_keys_for_comments:
            column_fields = [col.get('field') for col in column_defs if isinstance(col.get('field'), str)]
            if column_fields:
                comment_qs = PLComment.objects.filter(row_key__in=row_keys_for_comments, column_key__in=column_fields)
            else:
                comment_qs = PLComment.objects.filter(row_key__in=row_keys_for_comments)

            for comment in comment_qs.select_related('created_by'):
                key = f"{comment.row_key}||{comment.column_key}"
                entry = comment_summary.setdefault(key, {
                    'total': 0,
                    'open': 0,
                    'latest': None,
                })
                entry['total'] += 1
                if not comment.resolved:
                    entry['open'] += 1
                updated_iso = comment.updated_at.isoformat()
                if not entry['latest'] or updated_iso > entry['latest']:
                    entry['latest'] = updated_iso

    debug_info['ping'] = 'pl_report_data v4 - Fixed indexing and Decimal types'

    return JsonResponse({
        'columnDefs': column_defs,
        'rowData': row_data,
        'debug_info': debug_info,
        'commentSummary': comment_summary,
    })


def _serialize_pl_comment(comment, current_user=None):
    creator = comment.created_by
    display_name = creator.get_full_name() or creator.get_username()
    initials = ''.join([part[0] for part in (creator.get_full_name() or creator.get_username()).split() if part])[:2].upper()
    
    # Get files for this comment
    files = comment.files.all()
    serialized_files = [_serialize_pl_comment_file(file_obj, current_user) for file_obj in files] if current_user else []
    
    return {
        'id': comment.id,
        'parent_id': comment.parent_id,
        'row_key': comment.row_key,
        'column_key': comment.column_key,
        'row_label': comment.row_label,
        'column_label': comment.column_label,
        'message': comment.message,
        'resolved': comment.resolved,
        'created_at': comment.created_at.isoformat(),
        'updated_at': comment.updated_at.isoformat(),
        'created_by': {
            'id': creator.id,
            'name': display_name,
            'initials': initials or display_name[:2].upper(),
        },
        'can_edit': bool(current_user and (current_user.is_staff or creator_id_matches(current_user, creator))),
        'files': serialized_files,
    }


def creator_id_matches(current_user, creator):
    try:
        return current_user.id == creator.id
    except Exception:
        return False


def _comment_summary_for_cell(row_key, column_key):
    qs = PLComment.objects.filter(row_key=row_key, column_key=column_key)
    total = qs.count()
    open_count = qs.filter(resolved=False).count()
    latest = qs.order_by('-updated_at').values_list('updated_at', flat=True).first()
    return {
        'total': total,
        'open': open_count,
        'latest': latest.isoformat() if latest else None,
    }


@login_required
def pl_comment_list(request):
    if request.method == 'GET':
        row_key = request.GET.get('row_key')
        column_key = request.GET.get('column_key')
        if not row_key or not column_key:
            return JsonResponse({'error': 'row_key and column_key are required'}, status=400)
        comments = PLComment.objects.filter(row_key=row_key, column_key=column_key).select_related('created_by').order_by('created_at')
        data = [_serialize_pl_comment(comment, request.user) for comment in comments]
        summary = _comment_summary_for_cell(row_key, column_key) if comments else {'total': 0, 'open': 0, 'latest': None}
        return JsonResponse({'comments': data, 'summary': summary})

    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)

        message = (payload.get('message') or '').strip()
        row_key = payload.get('row_key')
        column_key = payload.get('column_key')
        if not message:
            return JsonResponse({'error': 'message is required'}, status=400)
        if not row_key or not column_key:
            return JsonResponse({'error': 'row_key and column_key are required'}, status=400)

        parent = None
        parent_id = payload.get('parent_id')
        if parent_id:
            parent = get_object_or_404(PLComment, pk=parent_id)

        comment = PLComment.objects.create(
            row_key=row_key,
            column_key=column_key,
            row_label=payload.get('row_label', ''),
            column_label=payload.get('column_label', ''),
            message=message,
            created_by=request.user,
            parent=parent,
        )

        summary = _comment_summary_for_cell(row_key, column_key)
        return JsonResponse({
            'comment': _serialize_pl_comment(comment, request.user),
            'summary': summary,
        }, status=201)

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def pl_comment_detail(request, pk):
    comment = get_object_or_404(PLComment, pk=pk)
    if request.method == 'PATCH':
        if not (request.user.is_staff or request.user == comment.created_by):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        try:
            payload = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)

        fields_to_update = []
        if 'message' in payload and isinstance(payload['message'], str):
            comment.message = payload['message'].strip()
            fields_to_update.append('message')
        if 'resolved' in payload:
            comment.resolved = bool(payload['resolved'])
            fields_to_update.append('resolved')

        if fields_to_update:
            fields_to_update.append('updated_at')
            comment.save(update_fields=fields_to_update)

        summary = _comment_summary_for_cell(comment.row_key, comment.column_key)
        return JsonResponse({
            'comment': _serialize_pl_comment(comment, request.user),
            'summary': summary,
        })

    if request.method == 'DELETE':
        if not (request.user.is_staff or request.user == comment.created_by):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        row_key = comment.row_key
        column_key = comment.column_key
        comment.delete()
        summary = _comment_summary_for_cell(row_key, column_key)
        return JsonResponse({'status': 'deleted', 'summary': summary})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


def _serialize_pl_comment_file(file_obj, user):
    """Serialize a PLCommentFile object for JSON response."""
    return {
        'id': file_obj.id,
        'original_filename': file_obj.original_filename,
        'file_size': file_obj.file_size,
        'file_size_human': file_obj.file_size_human,
        'file_type': file_obj.file_type,
        'uploaded_by': {
            'id': file_obj.uploaded_by.id,
            'username': file_obj.uploaded_by.username,
            'first_name': file_obj.uploaded_by.first_name,
            'last_name': file_obj.uploaded_by.last_name,
        },
        'uploaded_at': file_obj.uploaded_at.isoformat(),
        'download_url': file_obj.file.url if file_obj.file else None,
        'can_delete': user.is_staff or user == file_obj.uploaded_by,
    }


@login_required
def pl_comment_file_upload(request):
    """Upload a file to a P&L comment."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        comment_id = request.POST.get('comment_id')
        if not comment_id:
            return JsonResponse({'error': 'comment_id is required'}, status=400)
        
        comment = get_object_or_404(PLComment, pk=comment_id)
        
        # Check permissions
        if not (request.user.is_staff or request.user == comment.created_by):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        # Validate file type
        allowed_types = ['application/pdf', 'application/vnd.ms-excel', 
                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        'application/vnd.ms-excel.sheet.macroEnabled.12']
        if uploaded_file.content_type not in allowed_types:
            return JsonResponse({
                'error': 'Invalid file type. Only PDF and Excel files are allowed.'
            }, status=400)
        
        # Validate file size (max 10MB)
        max_size = 10 * 1024 * 1024  # 10MB
        if uploaded_file.size > max_size:
            return JsonResponse({
                'error': 'File too large. Maximum size is 10MB.'
            }, status=400)
        
        # Create the file record
        comment_file = PLCommentFile.objects.create(
            comment=comment,
            file=uploaded_file,
            original_filename=uploaded_file.name,
            file_size=uploaded_file.size,
            file_type=uploaded_file.content_type,
            uploaded_by=request.user
        )
        
        return JsonResponse({
            'file': _serialize_pl_comment_file(comment_file, request.user)
        }, status=201)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error uploading file: {error_details}")
        return JsonResponse({'error': str(e), 'details': error_details}, status=500)


@login_required
def pl_comment_file_delete(request, file_id):
    """Delete a file from a P&L comment."""
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    comment_file = get_object_or_404(PLCommentFile, pk=file_id)
    
    # Check permissions
    if not (request.user.is_staff or request.user == comment_file.uploaded_by):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    comment_file.delete()
    return JsonResponse({'status': 'deleted'})


@login_required
def bs_report_data(request):
    """Balance Sheet Report data in JSON format for AG Grid."""
    from_month = request.GET.get('from_month', '')
    from_year = request.GET.get('from_year', '')
    to_month = request.GET.get('to_month', '')
    to_year = request.GET.get('to_year', '')
    data_type_raw = request.GET.get('data_type', 'actual')
    data_type = data_type_raw.capitalize() if data_type_raw else 'Actual'
    
    from_date_start, from_date_end = convert_month_year_to_date_range(from_month, from_year)
    to_date_start, to_date_end = convert_month_year_to_date_range(to_month, to_year)
    
    # Get all companies (exclude pseudo-companies like Budget)
    companies = list(Company.objects.filter(is_budget_only=False).order_by('name'))
    company_codes = {c.code for c in companies}
    
    # Get ASSET, LIABILITY, EQUITY accounts from ChartOfAccounts
    bs_types = [
        'ASSET', 'LIABILITY', 'EQUITY',
        'Bank', 'Fixed Asset', 'Other Current Asset', 'Other Asset',
        'Other Current Liabilities', 'Other Current Liability',
        'Equity'
    ]
    q_objects = Q()
    for t in bs_types:
        q_objects |= Q(account_type__iexact=t)
    chart_accounts = list(ChartOfAccounts.objects.filter(q_objects).order_by('sort_order'))
    
    # If ChartOfAccounts is empty, return empty data
    if not chart_accounts:
        logger.warning("ChartOfAccounts is empty for Balance Sheet")
        return JsonResponse({
            'columnDefs': [],
            'rowData': []
        })
    
    # Get unique periods from FinancialData with proper filtering
    try:
        financial_data_query = FinancialData.objects.filter(data_type=data_type)
        
        if from_date_start:
            financial_data_query = financial_data_query.filter(period__gte=from_date_start)
        if to_date_end:
            financial_data_query = financial_data_query.filter(period__lte=to_date_end)
        
        periods = list(financial_data_query.values_list('period', flat=True).distinct().order_by('period'))
    except Exception as e:
        periods = []
    
    # If no periods, return empty data
    if not periods:
        return JsonResponse({
            'columnDefs': [],
            'rowData': []
        })
    
    # Pre-fetch all FinancialData for better performance
    financial_data = {}
    for period in periods:
        financial_data[period] = {}
        for company in companies:
            financial_data[period][company.code] = {}
    
    # Get all FinancialData in one query
    all_financial_data = FinancialData.objects.filter(
        data_type=data_type,
        period__in=periods
    ).select_related('company')
    
    # Organize financial data by period, company, and account
    for fd in all_financial_data:
        if fd.period in financial_data and fd.company.code in financial_data[fd.period]:
            financial_data[fd.period][fd.company.code][fd.account_code] = fd.amount

    # Track which company-period combinations have non-zero data
    non_zero_company_periods = set()
    for period, company_map in financial_data.items():
        period_key = period.strftime('%Y-%m')
        for company_code, accounts_map in company_map.items():
            if company_code not in company_codes:
                continue
            if any(amount for amount in accounts_map.values()):
                non_zero_company_periods.add((period_key, company_code))

    # Get companies that actually have data (same logic as P&L report)
    companies_with_data = list(set(fd.company for fd in all_financial_data))
    if not companies_with_data:
        logger.warning("No companies with data found for Balance Sheet, using all companies as fallback")
        companies_with_data = companies
    else:
        logger.info(f"Balance Sheet companies with data: {[c.code for c in companies_with_data]}")
    
    # Group accounts by account_type and sub_category from ChartOfAccounts
    grouped_data = {}
    
    for account in chart_accounts:
        # Use account_type and sub_category directly from ChartOfAccounts
        account_type = account.account_type or 'UNCATEGORIZED'
        sub_category = account.sub_category or 'UNCATEGORIZED'
        
        # Create account type if it doesn't exist
        if account_type not in grouped_data:
            grouped_data[account_type] = {}
        
        # Create sub category if it doesn't exist
        if sub_category not in grouped_data[account_type]:
            grouped_data[account_type][sub_category] = []
        
        # Add account to sub category
        grouped_data[account_type][sub_category].append(account)
    
    
    # Build report data with hierarchical structure
    report_data = []
    
    # Fixed order for Balance Sheet structure: ASSETS → LIABILITIES → EQUITY
    bs_structure = ['ASSET', 'LIABILITY', 'EQUITY']
    
    # Process each account type in the fixed order
    for account_type in bs_structure:
        if account_type not in grouped_data:
            continue
        
        # Add account type header
        # Convert singular to plural for display
        display_name = account_type
        if account_type == 'ASSET':
            display_name = 'ASSETS'
        elif account_type == 'LIABILITY':
            display_name = 'LIABILITIES'
        # EQUITY stays the same
        
        report_data.append({
            'type': 'parent_header',
            'account_name': display_name,
            'account_code': '',
            'periods': {},
            'grand_totals': {}
        })
        
        # Process sub categories
        for sub_category, accounts in grouped_data[account_type].items():
            # Add sub header
            report_data.append({
                'type': 'sub_header',
                'account_name': sub_category,
                'account_code': '',
                'periods': {},
                'grand_totals': {}
            })
            
            # Process individual accounts
            for account in accounts:
                row_data = {
                    'type': 'account',
                    'account_name': account.account_name,
                    'account_code': account.account_code,
                    'periods': {},
                    'grand_totals': {}
                }
                
                # Calculate period totals for each company
                for period in periods:
                    row_data['periods'][period] = {}
                    period_total = 0
                    
                    for company in companies:
                        amount = financial_data[period][company.code].get(account.account_code, 0)
                        # Convert to float for AG Grid
                        row_data['periods'][period][company.code] = float(amount or 0)
                        period_total += amount or 0
                    
                    row_data['periods'][period]['TOTAL'] = float(period_total or 0)
                
                # Calculate grand totals
                for company in companies:
                    grand_total = sum(
                        financial_data[period][company.code].get(account.account_code, 0)
                        for period in periods
                    )
                    row_data['grand_totals'][company.code] = float(grand_total or 0)
                
                # Calculate overall grand total
                overall_grand_total = sum(
                    sum(financial_data[period][company.code].get(account.account_code, 0) for company in companies)
                    for period in periods
                )
                row_data['grand_totals']['TOTAL'] = float(overall_grand_total or 0)
                
                report_data.append(row_data)
            
            # Add sub total
            sub_total_data = {
                'type': 'sub_total',
                'account_name': f'Total {sub_category}',
                'account_code': '',
                'periods': {},
                'grand_totals': {}
            }
            
            # Calculate sub totals
            for period in periods:
                sub_total_data['periods'][period] = {}
                period_total = 0
                
                for company in companies:
                    company_total = sum(
                        financial_data[period][company.code].get(account.account_code, 0)
                        for account in accounts
                    )
                    sub_total_data['periods'][period][company.code] = float(company_total or 0)
                    period_total += company_total or 0
                
                sub_total_data['periods'][period]['TOTAL'] = float(period_total or 0)
            
            # Calculate grand totals for sub category
            for company in companies:
                grand_total = sum(
                    sum(financial_data[period][company.code].get(account.account_code, 0) for account in accounts)
                    for period in periods
                )
                sub_total_data['grand_totals'][company.code] = float(grand_total or 0)
            
            # Calculate overall grand total for sub category
            overall_grand_total = sum(
                sum(sum(financial_data[period][company.code].get(account.account_code, 0) for account in accounts) for company in companies_with_data)
                for period in periods
            )
            sub_total_data['grand_totals']['TOTAL'] = float(overall_grand_total or 0)
            
            report_data.append(sub_total_data)
            
            # Add spacer row after sub total
            report_data.append({
                'type': 'spacer',
                'account_name': '',
                'account_code': '',
                'periods': {},
                'grand_totals': {}
            })
        
        # Add account type total
        # Convert singular to plural for display
        display_name = account_type
        if account_type == 'ASSET':
            display_name = 'ASSETS'
        elif account_type == 'LIABILITY':
            display_name = 'LIABILITIES'
        # EQUITY stays the same
        
        account_type_total_data = {
            'type': 'parent_total',
            'account_name': f'TOTAL {display_name}',
            'account_code': '',
            'periods': {},
            'grand_totals': {}
        }
        
        # Calculate account type totals
        for period in periods:
            account_type_total_data['periods'][period] = {}
            period_total = 0
            
            for company in companies:
                company_total = sum(
                    sum(financial_data[period][company.code].get(account.account_code, 0) for account in sub_accounts)
                    for sub_accounts in grouped_data[account_type].values()
                )
                account_type_total_data['periods'][period][company.code] = float(company_total or 0)
                period_total += company_total or 0
            
            account_type_total_data['periods'][period]['TOTAL'] = float(period_total or 0)
        
        # Calculate grand totals for account type
        for company in companies_with_data:
            grand_total = sum(
                sum(sum(financial_data[period][company.code].get(account.account_code, 0) for account in sub_accounts) for sub_accounts in grouped_data[account_type].values())
                for period in periods
            )
            account_type_total_data['grand_totals'][company.code] = float(grand_total or 0)
        
        # Calculate overall grand total for account type
        overall_grand_total = sum(
            sum(sum(sum(financial_data[period][company.code].get(account.account_code, 0) for account in sub_accounts) for sub_accounts in grouped_data[account_type].values()) for company in companies_with_data)
            for period in periods
        )
        account_type_total_data['grand_totals']['TOTAL'] = float(overall_grand_total or 0)
        
        report_data.append(account_type_total_data)
        
        # Add spacer row after account type total (except for the last one before CHECK)
        # Don't add spacer after TOTAL EQUITY since CHECK row follows immediately
        if account_type != 'EQUITY':
            report_data.append({
                'type': 'spacer',
                'account_name': '',
                'account_code': '',
                'periods': {},
                'grand_totals': {}
            })
    
    # Add CHECK row at bottom: TOTAL ASSETS - TOTAL LIABILITIES - TOTAL EQUITY (should equal 0)
    if 'ASSET' in grouped_data and ('LIABILITY' in grouped_data or 'EQUITY' in grouped_data):
        # Calculate totals for each category
        assets_total = 0
        liabilities_total = 0
        equity_total = 0
        
        if 'ASSET' in grouped_data:
            assets_total = sum(
                sum(sum(financial_data[period][company.code].get(account.account_code, 0) for account in sub_accounts) for sub_accounts in grouped_data['ASSET'].values())
                for period in periods
                for company in companies_with_data
            )
        
        if 'LIABILITY' in grouped_data:
            liabilities_total = sum(
                sum(sum(financial_data[period][company.code].get(account.account_code, 0) for account in sub_accounts) for sub_accounts in grouped_data['LIABILITY'].values())
                for period in periods
                for company in companies_with_data
            )
        
        if 'EQUITY' in grouped_data:
            equity_total = sum(
                sum(sum(financial_data[period][company.code].get(account.account_code, 0) for account in sub_accounts) for sub_accounts in grouped_data['EQUITY'].values())
                for period in periods
                for company in companies_with_data
            )
        
        check_value = assets_total - liabilities_total - equity_total
        
        report_data.append({
            'type': 'check_row',
            'account_name': 'CHECK (Assets - Liabilities - Equity)',
            'account_code': '',
            'periods': {period: {'TOTAL': float(check_value or 0)} for period in periods},
            'grand_totals': {'TOTAL': float(check_value or 0)}
        })
    
    # Prepare column definitions for AG Grid
    column_defs = [
        # The NEW toggler column
        {
            "headerName": "+", 
            "field": "toggler", # Can be any unique name
            "headerComponent": "togglerHeader", # We will create this in JavaScript
            "width": 40,
            "pinned": "left",
            "hide": False # This column is initially visible
        },
        # The MODIFIED A/C column
        {
            "field": "account_code", 
            "colId": "account_code", # Use colId for reliable access
            "headerName": "A/C",
            "headerComponent": "hideableHeader", # Custom header with a "-" button
            "width": 90, 
            "pinned": "left",
            "hide": True # This column is initially HIDDEN
        },
        {'field': 'account_name', 'headerName': 'Account Name', 'pinned': 'left', 'width': 250}
    ]
    
    # Add period columns
    for period in periods:
        for company in companies:
            column_defs.append({
                'field': f'{period.strftime("%b-%y")}_{company.code}',
                'headerName': f'{period.strftime("%b-%y")} {company.code}',
                'width': 120,
                'type': 'numberColumnWithCommas',
                'colType': 'company',
                'periodKey': period.strftime('%Y-%m'),
                'companyCode': company.code,
                'cellStyle': {
                    'textAlign': 'right',
                    # Styling based on company id instead of code prefix
                    'backgroundColor': '#E6F3FF' if getattr(company, 'id', None) == 1 else '#E8F5E9'
                }
            })
        column_defs.append({
            'field': f'{period.strftime("%b-%y")}_TOTAL',
            'headerName': f'{period.strftime("%b-%y")} TOTAL',
            'headerComponent': 'periodToggleHeader',
            'width': 120,
            'type': 'numberColumnWithCommas',
            'colType': 'total',
            'periodKey': period.strftime('%Y-%m'),
            'cellStyle': {
                'textAlign': 'right',
                'backgroundColor': '#FFF9E6'
            }
        })

    for col in column_defs:
        if col.get('colType') != 'company':
            continue
        period_key = col.get('periodKey')
        company_code = col.get('companyCode')
        if not period_key or not company_code:
            continue
        col['hide'] = (period_key, company_code) not in non_zero_company_periods

    
    # Prepare row data for AG Grid
    row_data = []
    for row in report_data:
        grid_row = {
            'account_code': row['account_code'],
            'account_name': row['account_name'],
            'rowType': row['type']
        }
        
        # Add period data
        for period in periods:
            for company in companies:
                field_name = f'{period.strftime("%b-%y")}_{company.code}'
                value = row['periods'].get(period, {}).get(company.code, 0)
                grid_row[field_name] = float(value or 0)
            
            field_name = f'{period.strftime("%b-%y")}_TOTAL'
            value = row['periods'].get(period, {}).get('TOTAL', 0)
            grid_row[field_name] = float(value or 0)
        
        # Apply row styling based on type
        if row['type'] == 'parent_header':
            grid_row['cellStyle'] = {'backgroundColor': '#e8f5e8', 'fontWeight': 'bold', 'fontSize': '14px'}
        elif row['type'] == 'sub_header':
            grid_row['cellStyle'] = {'backgroundColor': '#f0f8ff', 'fontWeight': 'bold'}
        elif row['type'] == 'sub_total':
            # Sub total rows - styling handled by JavaScript getRowClass
            grid_row['cellStyle'] = {
                'fontWeight': 'bold'
            }
        elif row['type'] == 'parent_total':
            # Parent total rows - styling handled by JavaScript getRowClass
            grid_row['cellStyle'] = {
                'fontWeight': 'bold'
            }
        elif row['type'] == 'check_row':
            grid_row['cellStyle'] = {'backgroundColor': '#ffe6e6', 'fontWeight': 'bold'}
        elif row['type'] == 'spacer':
            # Spacer rows - empty with minimal height
            grid_row['cellStyle'] = {
                'backgroundColor': 'transparent',
                'height': '10px',
                'border': 'none'
            }
        
        row_data.append(grid_row)
    
    # Filter out rows where ALL TOTAL columns sum to zero within selected date range
    # BUT preserve headers, totals, and check rows regardless of their values
    filtered_rows = []
    for row in row_data:
        row_type = row.get('rowType', 'unknown')
        
        # Always keep headers, totals, check rows, and spacer rows
        if row_type in ['parent_header', 'sub_header', 'sub_total', 'parent_total', 'check_row', 'spacer']:
            filtered_rows.append(row)
            continue
            
        # For account rows, check if they have non-zero totals
        total_sum = 0
        for period in periods:  # Only selected periods
            total_field = f'{period.strftime("%b-%y")}_TOTAL'
            total_sum += abs(float(row.get(total_field, 0)))
        
        if total_sum != 0:  # Keep account row if any TOTAL is non-zero
            filtered_rows.append(row)
    
    row_data = filtered_rows
    
    return JsonResponse({
        'columnDefs': column_defs,
        'rowData': row_data
    })

@login_required
def pl_report(request):
    """P&L Report view - simplified to only render template, data comes from JSON endpoint."""
    from_month = request.GET.get('from_month', '')
    from_year = request.GET.get('from_year', '')
    to_month = request.GET.get('to_month', '')
    to_year = request.GET.get('to_year', '')
    data_type_raw = request.GET.get('data_type', 'actual')
    data_type = data_type_raw.capitalize() if data_type_raw else 'Actual'
    
    year_range = range(2024, 2031)
    
    context = {
        'from_month': from_month,
        'from_year': from_year,
        'to_month': to_month,
        'to_year': to_year,
        'data_type': data_type,
        'report_type': 'pl',
        'year_range': year_range
    }
    
    return render(request, 'core/pl_report.html', context)

@login_required
def bs_report(request):
    """Balance Sheet Report view - simplified to only render template, data comes from JSON endpoint."""
    from_month = request.GET.get('from_month', '')
    from_year = request.GET.get('from_year', '')
    to_month = request.GET.get('to_month', '')
    to_year = request.GET.get('to_year', '')
    data_type_raw = request.GET.get('data_type', 'actual')
    data_type = data_type_raw.capitalize() if data_type_raw else 'Actual'

    year_range = range(2023, 2031)

    context = {
        'from_month': from_month,
        'from_year': from_year,
        'to_month': to_month,
        'to_year': to_year,
        'data_type': data_type,
        'report_type': 'bs',
        'year_range': year_range
    }

    return render(request, 'core/bs_report.html', context)

@login_required
def export_report_excel(request):
    """Export report data to Excel."""
    export_type = request.GET.get('export_type', 'raw')  # formatted | raw
    report_type = request.GET.get('type', 'pl')
    from_month = request.GET.get('from_month', '')
    from_year = request.GET.get('from_year', '')
    to_month = request.GET.get('to_month', '')
    to_year = request.GET.get('to_year', '')
    data_type_raw = request.GET.get('data_type', 'actual')
    data_type = data_type_raw.capitalize() if data_type_raw else 'Actual'
    
    from_date_start, from_date_end = convert_month_year_to_date_range(from_month, from_year)
    to_date_start, to_date_end = convert_month_year_to_date_range(to_month, to_year)
    
    companies = Company.objects.all().order_by('name')

    # Formatted export for P&L: reuse the same data structure as the screen
    if export_type == 'formatted' and report_type == 'pl':
        # Call pl_report_data to assemble hierarchical data
        screen_response = pl_report_data(request)
        try:
            screen_json = json.loads(screen_response.content.decode('utf-8'))
        except Exception:
            # Fallback: return empty workbook with a note
            df = pd.DataFrame([["No data"]], columns=["P&L Report"])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="pl_report_formatted.xlsx"'
            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='P&L Report', index=False)
            return response

        row_data = screen_json.get('rowData', [])
        column_defs = screen_json.get('columnDefs', [])

        # Build header: Account Name | Type | period columns (company, TOTAL, Budget) in the same order as the grid
        header = ['Account Name', 'Type']
        period_fields = []
        for col in column_defs:
            field = col.get('field')
            col_type = col.get('colType')
            if not field:
                continue
            # skip non-data fields
            if field in ('toggler', 'account_code', 'account_name'):
                continue
            # include only period data columns (company, total, budget)
            if col_type in ('company', 'total', 'budget', 'grand_company', 'grand_overall', 'grand_budget'):
                header.append(col.get('headerName', field))
                period_fields.append(field)

        # Fallback if no column defs were found (use keys from first row excluding meta keys)
        if not period_fields and row_data:
            meta_keys = {'account_name', 'account_code', 'rowType', 'cellStyle'}
            first_row = row_data[0]
            for k in first_row.keys():
                if k not in meta_keys:
                    header.append(k)
                    period_fields.append(k)

        # Assemble rows
        excel_rows = []
        for r in row_data:
            name = r.get('account_name', '')
            rtype = r.get('rowType', '')
            values = []
            for f in period_fields:
                v = r.get(f, 0)
                if v == '-':
                    v = 0
                try:
                    v = float(v)
                except Exception:
                    v = 0
                values.append(v)
            excel_rows.append([name, rtype] + values)

        # Create DataFrame and export
        df = pd.DataFrame(excel_rows, columns=header)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pl_report_formatted.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            sheet_name = 'P&L Report'
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply basic formatting: bold for subtotal/total/header rows, light fill for totals
            ws = writer.sheets.get(sheet_name)
            if ws is not None:
                # Map row types we want bolded/fill
                bold_types = {'sub_total', 'parent_total', 'total', 'section_header', 'parent_header', 'net_income'}
                total_fill = openpyxl.styles.PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
                bold_font = openpyxl.styles.Font(bold=True)

                # Iterate DataFrame rows (1-based Excel, header at row 1)
                type_col_idx = header.index('Type') + 1  # 1-based
                max_col = ws.max_column
                for row_idx in range(2, ws.max_row + 1):
                    row_type_val = ws.cell(row=row_idx, column=type_col_idx).value
                    if row_type_val in bold_types:
                        for col_idx in range(1, max_col + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.font = bold_font
                            if row_type_val == 'total':
                                cell.fill = total_fill

        return response
    
    if report_type == 'pl':
        accounts = ChartOfAccounts.objects.filter(
            account_type__in=['INCOME', 'EXPENSE']
        ).order_by('sort_order')
        report_title = 'Profit & Loss Report'
    else:
        accounts = ChartOfAccounts.objects.filter(
            account_type__in=['ASSET', 'LIABILITY', 'EQUITY']
        ).order_by('sort_order')
        report_title = 'Balance Sheet Report'
    
    account_codes = accounts.values_list('account_code', flat=True)
    periods_query = FinancialData.objects.filter(account_code__in=account_codes)
    
    if from_date_start:
        periods_query = periods_query.filter(period__gte=from_date_start)
    if to_date_end:
        periods_query = periods_query.filter(period__lte=to_date_end)
    
    periods = periods_query.values_list('period', flat=True).distinct().order_by('period')
    
    # Prepare Excel data
    excel_data = []
    
    # Add headers
    header_row = ['Account Code', 'Account Name']
    for period in periods:
        for company in companies:
            header_row.append(f"{period.strftime('%Y-%m')} {company.code}")
        header_row.append(f"{period.strftime('%Y-%m')} TOTAL")
    excel_data.append(header_row)
    
    # Add data rows
    for account in accounts:
        row = [account.account_code or '', account.account_name]
        
        for period in periods:
            period_total = 0
            for company in companies:
                amount = FinancialData.objects.filter(
                    company=company,
                    account_code=account.account_code,
                    period=period,
                    data_type=data_type
                ).aggregate(total=Sum('amount'))['total'] or 0
                row.append(float(amount))
                period_total += amount
            row.append(float(period_total))
        
        excel_data.append(row)
    
    # Create DataFrame and export
    df = pd.DataFrame(excel_data[1:], columns=excel_data[0])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_title.lower().replace(" ", "_")}.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=report_title, index=False)
    
    return response


@login_required
def export_for_stakeholders(request):
    import json
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from decimal import Decimal
    
    screen_response = pl_report_data(request)
    try:
        screen_json = json.loads(screen_response.content.decode('utf-8'))
    except Exception:
        df = pd.DataFrame([["No data"]], columns=["P&L Report"])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pl_report_stakeholders.xlsx"'
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='P&L Report', index=False)
        return response
    
    row_data = screen_json.get('rowData', [])
    column_defs = screen_json.get('columnDefs', [])
    
    row_data = [
        row for row in row_data 
        if not row.get('is_cf_dashboard') 
        and not row.get('is_separator')
        and row.get('rowType') != 'cf_metric'
        and row.get('section') != 'cf_dashboard'
    ]
    
    if not row_data or not column_defs:
        df = pd.DataFrame([["No data available"]], columns=["P&L Report"])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="pl_report_stakeholders.xlsx"'
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='P&L Report', index=False)
        return response
    
    periods = []
    companies = []
    period_company_map = {}
    grand_total_fields = []
    
    for col in column_defs:
        field = col.get('field')
        col_type = col.get('colType')
        header_name = col.get('headerName', '')
        hide_flag = col.get('hide', False)
        
        if field in ('toggler', 'account_code', 'account_name'):
            continue
        
        if hide_flag:
            continue
        
        if col_type in ('company', 'total', 'budget'):
            parts = field.split('_')
            if len(parts) >= 2:
                period = parts[0]
                company_or_type = '_'.join(parts[1:])
                
                if period not in period_company_map:
                    period_company_map[period] = []
                    periods.append(period)
                
                period_company_map[period].append({
                    'field': field,
                    'name': header_name,
                    'type': col_type
                })
                
                if company_or_type not in companies and col_type == 'company':
                    companies.append(company_or_type)
        
        elif col_type in ('grand_company', 'grand_overall', 'grand_budget'):
            if hide_flag:
                continue
            grand_total_fields.append({
                'field': field,
                'name': header_name,
                'type': col_type
            })
            parts = field.split('_')
            if len(parts) >= 3:
                company_code = '_'.join(parts[2:])
                if company_code not in companies and col_type == 'grand_company':
                    companies.append(company_code)
    
    header_row1 = ['Account Name']
    header_row2 = ['']
    
    for period in periods:
        period_cols = period_company_map.get(period, [])
        num_cols = len(period_cols)
        header_row1.append(period)
        for i in range(num_cols - 1):
            header_row1.append('')
        
        for col_info in period_cols:
            header_row2.append(col_info['name'])
    
    if grand_total_fields:
        num_grand_cols = len(grand_total_fields)
        header_row1.append('Grand Total')
        for i in range(num_grand_cols - 1):
            header_row1.append('')
        
        for col_info in grand_total_fields:
            header_row2.append(col_info['name'])
    
    excel_rows = []
    for row in row_data:
        account_name = row.get('account_name', '')
        row_type = row.get('rowType', '')
        
        excel_row = [account_name]
        
        for period in periods:
            period_cols = period_company_map.get(period, [])
            for col_info in period_cols:
                field = col_info['field']
                value = row.get(field, 0)
                if value == '-':
                    value = ''
                elif value is not None:
                    try:
                        value = float(value)
                    except:
                        value = ''
                excel_row.append(value)
        
        for col_info in grand_total_fields:
            field = col_info['field']
            value = row.get(field, 0)
            if value == '-':
                value = ''
            elif value is not None:
                try:
                    value = float(value)
                except:
                    value = ''
            excel_row.append(value)
        
        excel_rows.append(excel_row)
    
    all_data = [header_row2] + excel_rows
    df = pd.DataFrame(all_data[1:], columns=all_data[0])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pl_report_stakeholders.xlsx"'
    
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='P&L Report', index=False, startrow=1)
        ws = writer.sheets['P&L Report']
        
        for col_idx, value in enumerate(header_row1, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = value
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        current_col = 2
        for period in periods:
            period_cols = period_company_map.get(period, [])
            num_cols = len(period_cols)
            
            if num_cols > 1:
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + num_cols - 1)
            
            current_col += num_cols
        
        if grand_total_fields:
            num_grand = len(grand_total_fields)
            if num_grand > 1:
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + num_grand - 1)
        
        for col_idx in range(1, len(header_row2) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        bold_types = {'sub_total', 'parent_total', 'total', 'section_header', 'parent_header', 'net_income'}
        total_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
        bold_font = Font(bold=True)
        
        for row_idx in range(3, ws.max_row + 1):
            excel_row_idx = row_idx - 3
            if excel_row_idx < len(excel_rows):
                source_row = row_data[excel_row_idx]
                row_type = source_row.get('rowType', '')
                
                if row_type in bold_types:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = bold_font
                        if row_type == 'total':
                            cell.fill = total_fill
        
        excel_col = 2
        for period in periods:
            period_cols = period_company_map.get(period, [])
            
            for i, col_info in enumerate(period_cols):
                col_letter = get_column_letter(excel_col + i)
                
                if col_info['type'] == 'company':
                    ws.column_dimensions[col_letter].outline_level = 1
            
            excel_col += len(period_cols)
        
        if grand_total_fields:
            for i, col_info in enumerate(grand_total_fields):
                col_letter = get_column_letter(excel_col + i)
                
                if col_info['type'] == 'grand_company':
                    ws.column_dimensions[col_letter].outline_level = 1
        
        for row_idx in range(3, ws.max_row + 1):
            excel_row_idx = row_idx - 3
            if excel_row_idx < len(excel_rows):
                source_row = row_data[excel_row_idx]
                row_type = source_row.get('rowType', '')
                
                if row_type == 'account':
                    ws.row_dimensions[row_idx].outline_level = 1
        
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.sheet_properties.outlinePr.summaryRight = True
        
        ws.column_dimensions['A'].width = 35
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        ws.freeze_panes = 'B3'
    
    return response


@csrf_exempt
@login_required
def update_cf_dashboard(request):
    """API endpoint for updating CF Dashboard values from inline editing"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            # Handle consolidated Budget/Forecast updates (no company dimension)
            if data.get('is_budget'):
                from .models import CFDashboardBudget
                # Parse period string similar to below
                period_str = data.get('period')
                if not period_str:
                    return JsonResponse({'status': 'error', 'message': 'Missing period'}, status=400)
                # Accept formats: "Jan-24" or "202401"
                if '-' in period_str:
                    month_map = {
                        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
                    }
                    month_str, year_str = period_str.split('-')
                    month = month_map.get(month_str, 1)
                    year = 2000 + int(year_str) if len(year_str) == 2 else int(year_str)
                else:
                    year = int(period_str[:4])
                    month = int(period_str[4:6])
                period_date = date(year, month, 1)

                metric_id = data.get('metric_id')
                if not metric_id:
                    return JsonResponse({'status': 'error', 'message': 'Missing metric_id'}, status=400)
                # Determine type (default to budget)
                dtype = data.get('data_type')
                if dtype not in ('budget', 'forecast'):
                    dtype = 'budget'

                CFDashboardBudget.objects.update_or_create(
                    metric_id=metric_id,
                    period=period_date,
                    data_type=dtype,
                    defaults={'value': data.get('value', 0)}
                )
                return JsonResponse({'status': 'success'})
            
            # Parse period from different formats
            period_str = data['period']
            print(f"Period string received: {period_str}")
            
            # Handle format "Jan-24" or "202401"
            if '-' in period_str:
                # Format: "Jan-24"
                month_map = {
                    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
                }
                month_str, year_str = period_str.split('-')
                month = month_map.get(month_str, 1)
                year = 2000 + int(year_str) if len(year_str) == 2 else int(year_str)
            else:
                # Format: "202401"
                year = int(period_str[:4])
                month = int(period_str[4:6])
            
            period_date = date(year, month, 1)
            print(f"Parsed date: {period_date}")
            
            # Get company by code
            company = Company.objects.get(code__iexact=data['company_code'])
            print(f"Found company: {company}")
            
            # Get metric
            metric = CFDashboardMetric.objects.get(id=data['metric_id'])
            print(f"Found metric: {metric}")
            
            # Update or create the data
            cf_data, created = CFDashboardData.objects.update_or_create(
                company=company,
                period=period_date,
                metric=metric,
                defaults={'value': data['value']}
            )
            print(f"Data saved: {cf_data}, created: {created}")
            
            return JsonResponse({'status': 'success', 'created': created})
            
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)


@login_required
@permission_required('core.view_salary_details')  
def salary_details(request):
    from django.db.models import Sum
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    import calendar

    company_code = request.GET.get('company')
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not all([company_code, year, month]):
        return HttpResponse("Missing parameters", status=400)

    try:
        year_int = int(year)
        month_int = int(month)
    except (TypeError, ValueError):
        return HttpResponse("Invalid year or month", status=400)

    show_all_companies = company_code.upper() == 'ALL'

    base_queryset = SalaryData.objects.select_related('company').filter(
        year=year_int,
        month=month_int
    )

    if show_all_companies:
        salaries = base_queryset.order_by('company__code', 'employee_name')
        company = None
        company_label = "All Companies"
    else:
        company = get_object_or_404(Company, code=company_code)
        salaries = base_queryset.filter(company=company).order_by('employee_name')
        company_label = company.name

    month_name = calendar.month_name[month_int] if 1 <= month_int <= 12 else str(month_int)

    context = {
        'company': company,
        'company_label': company_label,
        'company_code': company_code,
        'year': year_int,
        'month': month_int,
        'month_name': month_name,
        'salaries': salaries,
        'total': salaries.aggregate(Sum('amount'))['amount__sum'] or 0,
        'show_all_companies': show_all_companies
    }

    return render(request, 'core/salary_details.html', context)


@login_required
@permission_required('core.upload_salary_data')
def upload_salaries(request):
    import calendar
    from decimal import Decimal
    from django.db import transaction
    
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        try:
            df = pd.read_csv(file)
            
            # Check all companies exist first
            company_codes = df['Ent'].unique()
            missing = []
            for code in company_codes:
                if not Company.objects.filter(code=code).exists():
                    missing.append(code)
            
            if missing:
                messages.error(request, f"Companies not found: {', '.join(missing)}")
                return redirect('/salaries/upload/')
            
            # Process in transaction
            with transaction.atomic():
                # Month header helpers so we support "Jan-25" AND "25-Jan" formats
                month_abbr_map = {
                    abbr.lower(): idx for idx, abbr in enumerate(calendar.month_abbr) if abbr
                }
                month_name_map = {
                    name.lower(): idx for idx, name in enumerate(calendar.month_name) if name
                }

                def parse_month_year(raw_header: str) -> tuple[int, int]:
                    parts = [p.strip() for p in raw_header.split('-') if p.strip()]
                    if len(parts) != 2:
                        raise ValueError(f"Column '{raw_header}' does not look like a month header")

                    def detect_month(token: str) -> int | None:
                        key = token.lower()
                        if key in month_abbr_map:
                            return month_abbr_map[key]
                        if key in month_name_map:
                            return month_name_map[key]
                        return None

                    def detect_year(token: str) -> int | None:
                        if not token or not token.replace(' ', '').replace("'", '').replace('`', '').isdigit():
                            return None
                        cleaned = ''.join(ch for ch in token if ch.isdigit())
                        if len(cleaned) == 2:
                            return 2000 + int(cleaned)
                        return int(cleaned)

                    month_val = detect_month(parts[0])
                    year_val = detect_year(parts[1])
                    if month_val and year_val:
                        return month_val, year_val

                    # Try reversed order (e.g. "25-Jan")
                    month_val = detect_month(parts[1])
                    year_val = detect_year(parts[0])
                    if month_val and year_val:
                        return month_val, year_val

                    raise ValueError(f"Could not parse month/year from header '{raw_header}'")

                # Get month columns (Jan-25, Feb-25, etc)
                month_cols = [col for col in df.columns if '-' in str(col)]

                for _, row in df.iterrows():
                    company_code = str(row['Ent']).strip()
                    company = Company.objects.get(code=company_code)

                    for month_col in month_cols:
                        cell_value = row[month_col]
                        if pd.isna(cell_value) or str(cell_value).strip() == '':
                            continue

                        try:
                            month_num, year = parse_month_year(str(month_col))
                        except ValueError as exc:
                            raise ValueError(f"{exc} for employee '{row['Employee Name']}'") from exc

                        # Clean amount (keep digits, dot and minus)
                        amount_str = ''.join(ch for ch in str(cell_value) if ch.isdigit() or ch in ['.', '-', ','])
                        amount_str = amount_str.replace(',', '')
                        if amount_str in ['', '-', '.']:
                            continue
                        try:
                            amount = Decimal(amount_str)
                        except Exception as exc:
                            raise ValueError(
                                f"Unable to parse amount '{cell_value}' for {month_col} ({row['Employee Name']})"
                            ) from exc

                        employee_id = str(row['Employee ID']).strip() if pd.notna(row['Employee ID']) else ''

                        # Create or update
                        SalaryData.objects.update_or_create(
                            employee_id=employee_id,
                            company=company,
                            month=month_num,
                            year=year,
                            defaults={
                                'employee_name': str(row['Employee Name']).strip(),
                                'amount': amount,
                                'uploaded_by': request.user
                            }
                        )
            
            messages.success(request, 'Salaries uploaded successfully')
            return redirect('upload_salaries')
            
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('/salaries/upload/')
    
    return render(request, 'core/upload_salaries.html')


@login_required
def download_salary_template(request):
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="salary_template.csv"'

    writer = csv.writer(response)
    # Header row
    writer.writerow(['Ent', 'Employee ID', 'Employee Name', 'Jan-25', 'Feb-25', 'Mar-25', 'Apr-25', 'May-25', 'Jun-25', 'Jul-25', 'Aug-25', 'Sep-25', 'Oct-25', 'Nov-25', 'Dec-25'])
    # Example rows
    writer.writerow(['FG', 'EMP001', 'John Doe', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00', '$5000.00'])
    writer.writerow(['F2', 'EMP002', 'Jane Smith', '$4500.00', '$4500.00', '$4500.00', '', '', '', '', '', '', '', '', ''])

    return response


@login_required
@require_http_methods(["POST"])
def hubspot_sync(request):
    """Trigger HubSpot CRM synchronization for the requested object types."""

    raw_objects = request.POST.get('objects') or request.GET.get('objects')
    requested = [item.strip().lower() for item in raw_objects.split(',')] if raw_objects else []

    allowed = {'deals', 'companies', 'contacts', 'all'}
    invalid = [item for item in requested if item and item not in allowed]
    if invalid:
        return JsonResponse({'error': f"Unsupported object types requested: {', '.join(invalid)}"}, status=400)

    objects_to_sync = ['deals', 'companies', 'contacts'] if not requested or 'all' in requested else []
    if not objects_to_sync:
        for item in requested:
            if item not in objects_to_sync and item in allowed:
                objects_to_sync.append(item)

    try:
        service = HubSpotService()
    except ImproperlyConfigured as exc:
        logger.warning("HubSpot sync attempted without access token configured")
        return JsonResponse({'error': str(exc)}, status=500)

    action_map = {
        'deals': service.sync_deals,
        'companies': service.sync_companies,
        'contacts': service.sync_contacts,
    }

    results = {}
    for object_type in objects_to_sync:
        try:
            results[object_type] = action_map[object_type]()
        except Exception as exc:  # noqa: BLE001 - we want full traceback logged
            logger.exception("HubSpot %s sync failed", object_type)
            results[object_type] = {
                'status': HubSpotSyncLog.Status.FAILURE.value,
                'error': str(exc),
            }

    statuses = {result.get('status') for result in results.values()}
    if HubSpotSyncLog.Status.FAILURE.value in statuses:
        status_code = 500
    elif HubSpotSyncLog.Status.PARTIAL.value in statuses:
        status_code = 207
    else:
        status_code = 200

    metrics = service.get_financial_metrics()

    return JsonResponse(
        {
            'objects': objects_to_sync,
            'results': results,
            'metrics': metrics,
        },
        status=status_code,
    )


@login_required
@require_http_methods(["GET"])
def hubspot_data(request):
    """Return synchronized HubSpot data stored locally along with basic metrics."""

    record_type = request.GET.get('record_type')
    if record_type:
        record_type = record_type.strip().lower()
        if record_type not in HubSpotData.RecordType.values:
            return JsonResponse({'error': 'Invalid record_type provided.'}, status=400)

    limit_param = request.GET.get('limit')
    try:
        limit = int(limit_param) if limit_param else 100
    except (TypeError, ValueError):
        return JsonResponse({'error': 'limit parameter must be an integer.'}, status=400)

    limit = max(1, min(limit, 500))

    queryset = HubSpotData.objects.all()
    if record_type:
        queryset = queryset.filter(record_type=record_type)

    queryset = queryset.order_by('-synced_at')
    total_count = queryset.count()

    entries = [
        {
            'hubspot_id': item.hubspot_id,
            'record_type': item.record_type,
            'created_at': item.created_at.isoformat(),
            'synced_at': item.synced_at.isoformat(),
            'data': item.data,
        }
        for item in queryset[:limit]
    ]

    recent_logs = [
        {
            'id': log.id,
            'sync_type': log.sync_type,
            'status': log.status,
            'started_at': log.started_at.isoformat() if log.started_at else None,
            'finished_at': log.finished_at.isoformat() if log.finished_at else None,
            'details': log.details,
            'error_message': log.error_message,
        }
        for log in HubSpotSyncLog.objects.order_by('-started_at')[:10]
    ]

    try:
        metrics = HubSpotService().get_financial_metrics()
    except Exception as exc:  # noqa: BLE001 - ensure API errors are surfaced gracefully
        logger.exception("Failed to compute HubSpot financial metrics")
        metrics = {'error': str(exc)}

    return JsonResponse(
        {
            'record_type': record_type,
            'limit': limit,
            'total_count': total_count,
            'returned': len(entries),
            'results': entries,
            'recent_logs': recent_logs,
            'metrics': metrics,
        }
    )
